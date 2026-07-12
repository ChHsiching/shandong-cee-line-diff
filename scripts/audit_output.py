"""Slice C — data-quality audit hard gate (spec V5-3, Plan v2 Slice C 修订).

The completion gate for the admission-data pipeline. Before the pipeline is
declared "done", the **real produced xlsx** (not just pytest) must pass five
checks. The audit reads the produced xlsx + the verification jsonl and returns
an :class:`AuditReport`; :func:`main` exits 0 iff ``ok=True``.

iteration-3 (structured-columns, Plan v2 CRITICAL扩范围): every check that
previously inspected the legacy「匹配日志」cell now keys off the structured
「匹配阶段」column **by name** (not index, not the log string). This removes
the duplicated JUDGMENTAL_LOG_PREFIXES copy and survives column re-ordering.

Checks:
  0  judgmental_coverage — every row whose 匹配阶段 ∈ {核心名匹配, agent 语义匹配} in
     the hierarchical output must appear in ``verify_*_result.jsonl`` with
     verdict=确定. Missing jsonl entirely → fail with「复核未派发」.
  1  nonempty_stage — every 本科 major row in the flat output carries a
     non-empty 匹配阶段 (0 缺失).
  2  no_empty_rows — every produced table has zero fully-blank data rows.
  3  tables_nonempty — every produced table has at least one data row
     (field-mapping regression guard).
  4  jt_consistency — random ≥30 matched rows' J/T agree with the source
     近三年 values (matched rows) or with the estimate table (新增 rows).
     Matched = 匹配阶段 ∈ {严格匹配, 核心名匹配, agent 语义匹配}; estimate = 匹配阶段
     == 新增专业.

A side artefact ``audit_sample.xlsx`` is written for human semantic review; it
does NOT influence ``ok`` / the exit code.
"""

from __future__ import annotations

import argparse
import json
import logging
import random
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Sequence

import openpyxl

from scripts.stage0_merge import build_unified_history
from scripts.structured_log import JUDGMENTAL_STAGES
from scripts.write_outputs import HEADER_STAGE

__all__ = [
    "AuditReport",
    "AuditCheck",
    "HIER_ARCHICAL_NAME",
    "FLAT_NAME",
    "JUDGMENTAL_STAGES",
    "MATCHED_STAGES",
    "ESTIMATE_STAGE",
    "audit",
    "main",
]

logger = logging.getLogger(__name__)

# Produced-file names (mirror write_outputs / write_edge_tables).
HIER_ARCHICAL_NAME = "大绿本_完整版_含线差.xlsx"
FLAT_NAME = "大绿本_专业列表_含线差.xlsx"
NEW_MAJOR_NAME = "今年新增往年没有的专业.xlsx"
SPECIAL_NAME = "未能匹配的专业.xlsx"
DELETED_NAME = "往年有但今年停招的专业.xlsx"
RENAME_NAME = "学校改名表.xlsx"
NEW_SCHOOL_NAME = "今年新招生的学校.xlsx"
GONE_SCHOOL_NAME = "往年有今年停招的学校.xlsx"

# Every produced table the audit must consider for the empty-row / nonempty
# checks. Order is stable for readable reports.
PRODUCED_TABLES: tuple[str, ...] = (
    HIER_ARCHICAL_NAME,
    FLAT_NAME,
    NEW_MAJOR_NAME,
    SPECIAL_NAME,
    DELETED_NAME,
    RENAME_NAME,
    NEW_SCHOOL_NAME,
    GONE_SCHOOL_NAME,
)

# iteration-3: stage-based whitelists (replace the old JUDGMENTAL_LOG_PREFIXES
# copy). The audit reads 匹配阶段 by column NAME, so a future wording tweak in
# the log constants cannot silently drop a row from a check.
STAGE_HEADER = HEADER_STAGE  # #18d: 列名单点（write_outputs.HEADER_STAGE）
# Matched = 严格匹配 + 核心名匹配(Stage1.5 past=1，构造确定) + agent 语义匹配。
# 核心名匹配不是 JUDGMENTAL（豁免复核），但仍是 matched（jt_consistency 仍抽查）。
MATCHED_STAGES: frozenset[str] = JUDGMENTAL_STAGES | {"严格匹配", "核心名匹配"}
ESTIMATE_STAGE = "新增专业"

# Column indices in the hierarchical / flat output (1-based) — used only for
# the fixed-position columns (school / major / J / T / code). The structured
# columns are looked up BY NAME.
COL_SCHOOL = 4
COL_MAJOR_NAME = 6
COL_J = 13
COL_T = 14
# Major-row detector columns (1-based): 代号(E=5) + 名称(F=6) both non-empty.
COL_CODE = 5

# Sample size for the human-review artefact (Plan v2: 随机 ≥30).
SAMPLE_SIZE = 30
# Tolerance for matched-row J/T comparison (history values are exact; the
# output may carry the same float, so a tiny epsilon guards float drift).
JT_TOLERANCE = 0.011


# One check result: ``{name, passed, detail}`` (Plan v2 binding structure).
AuditCheck = dict[str, Any]


@dataclass
class AuditReport:
    """Aggregate audit outcome. ``ok`` is True iff every check passed.

    ``checks`` is a list of ``{name, passed, detail}`` dicts (Plan v2 binding
    structure) so callers can index by key.
    """

    ok: bool
    checks: list[AuditCheck] = field(default_factory=list)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _load_rows(path: Path) -> list[tuple]:
    """Read all rows of a workbook's active sheet as tuples."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()


def _is_major_row(row: tuple) -> bool:
    """A 本科 专业行: 代号 + 名称 both non-empty (mirrors write_outputs)."""
    if len(row) < COL_MAJOR_NAME:
        return False
    code = row[COL_CODE - 1]
    name = row[COL_MAJOR_NAME - 1]
    return code not in (None, "") and name not in (None, "")


def _stage_col_index(header: tuple) -> int | None:
    """Return the 0-based index of the 匹配阶段 column, or None if absent."""
    for i, h in enumerate(header):
        if h == STAGE_HEADER:
            return i
    return None


def _stage_of(row: tuple, stage_idx: int | None) -> str:
    """Return the row's 匹配阶段 value ("" when column absent or cell empty)."""
    if stage_idx is None or stage_idx >= len(row):
        return ""
    v = row[stage_idx]
    return "" if v is None else str(v)


def _load_verify_verdicts(semantic_dir: Path) -> dict[int, str] | None:
    """Read every ``verify_*_result.jsonl`` under semantic_dir.

    Returns ``{src_row_idx: verdict}`` or ``None`` when no verify result file
    exists at all (signals「复核未派发」to check 0).
    """
    paths = sorted(semantic_dir.glob("verify_*_result.jsonl"))
    if not paths:
        return None
    out: dict[int, str] = {}
    for p in paths:
        for raw in p.read_text(encoding="utf-8").splitlines():
            raw = raw.strip()
            if not raw:
                continue
            obj = json.loads(raw)
            idx = obj["src_row_idx"]
            out[idx] = obj["verdict"]
    return out


def _history_index(
    history: Sequence[dict[str, Any]],
) -> dict[tuple[str, str], list[dict[str, Any]]]:
    """Index history rows by (school, major) → list of rows (Bug #3 fix:
    same school+major can carry multiple 招生类别 with different J/T)."""
    idx: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for h in history:
        key = (h.get("school", ""), h.get("major", ""))
        idx.setdefault(key, []).append(h)
    return idx


def _row_fully_blank(row: tuple) -> bool:
    """True iff every cell is None or empty string (a '全空数据行')."""
    return all(c is None or c == "" for c in row)


# ---------------------------------------------------------------------------
# checks
# ---------------------------------------------------------------------------


def _check0_judgmental_coverage(
    hier_rows: list[tuple], verdicts: dict[int, str] | None
) -> AuditCheck:
    """Every row with 匹配阶段 ∈ {核心名匹配, agent 语义匹配} must have verdict=确定
    in verify_*_result.jsonl."""
    name = "judgmental_coverage"
    if verdicts is None:
        return {
            "name": name,
            "passed": False,
            "detail": "复核未派发：semantic-match/verify_*_result.jsonl 缺失",
        }

    if not hier_rows:
        return {"name": name, "passed": True, "detail": "无数据行"}

    header = tuple(h if h is not None else "" for h in hier_rows[0])
    stage_idx = _stage_col_index(header)

    missing: list[int] = []
    wrong_verdict: list[int] = []
    for row_idx_0, row in enumerate(hier_rows):
        src_row_idx = row_idx_0 + 1  # 1-based; header is row 1
        if src_row_idx == 1:
            continue
        if not _is_major_row(tuple(row)):
            continue
        stage = _stage_of(tuple(row), stage_idx)
        if stage not in JUDGMENTAL_STAGES:
            continue
        v = verdicts.get(src_row_idx)
        if v is None:
            missing.append(src_row_idx)
        elif v != "确定":
            wrong_verdict.append(src_row_idx)

    if missing or wrong_verdict:
        detail = (
            f"判断型匹配 {len(missing)} 行无复核结果、"
            f"{len(wrong_verdict)} 行 verdict≠确定；"
            f"示例缺 idx={missing[:5]}，verdict≠确定 idx={wrong_verdict[:5]}"
        )
        return {"name": name, "passed": False, "detail": detail}
    return {
        "name": name,
        "passed": True,
        "detail": "判断型匹配复核覆盖完备（verdict=确定）",
    }


def _check1_nonempty_stage(flat_rows: list[tuple]) -> AuditCheck:
    """Every 本科 major row in the flat output has a non-empty 匹配阶段."""
    name = "nonempty_log"  # keep historical check name for report stability
    if not flat_rows:
        return {"name": name, "passed": False, "detail": "扁平版无数据行"}
    header = tuple(h if h is not None else "" for h in flat_rows[0])
    stage_idx = _stage_col_index(header)

    blank: list[int] = []
    for row_idx_0, row in enumerate(flat_rows):
        if row_idx_0 == 0:
            continue  # header
        if not _is_major_row(tuple(row)):
            continue
        stage = _stage_of(tuple(row), stage_idx)
        if stage.strip() == "":
            blank.append(row_idx_0 + 1)
    if blank:
        return {
            "name": name,
            "passed": False,
            "detail": f"扁平版 {len(blank)} 行匹配阶段为空；示例行={blank[:5]}",
        }
    return {"name": name, "passed": True, "detail": "本科专业行匹配阶段全部非空"}


def _check2_no_empty_rows(out_dir: Path) -> AuditCheck:
    """No produced table contains a fully-blank data row."""
    name = "no_empty_rows"
    offenders: list[str] = []
    for fname in PRODUCED_TABLES:
        path = out_dir / fname
        if not path.exists():
            continue
        rows = _load_rows(path)
        for row_idx_0, row in enumerate(rows):
            if row_idx_0 == 0:
                continue  # header
            if _row_fully_blank(tuple(row)):
                offenders.append(f"{fname}:行{row_idx_0 + 1}")
                break  # one offender per table is enough
    if offenders:
        return {
            "name": name,
            "passed": False,
            "detail": f"{len(offenders)} 张表含全空数据行：{offenders[:5]}",
        }
    return {"name": name, "passed": True, "detail": "所有产出表无全空数据行"}


def _check3_tables_nonempty(out_dir: Path) -> AuditCheck:
    """Every produced table has at least one data row (header-only = regression)."""
    name = "tables_nonempty"
    empty_tables: list[str] = []
    for fname in PRODUCED_TABLES:
        path = out_dir / fname
        if not path.exists():
            empty_tables.append(f"{fname}(缺失)")
            continue
        rows = _load_rows(path)
        data_rows = [r for i, r in enumerate(rows) if i > 0]
        if not data_rows:
            empty_tables.append(fname)
    if empty_tables:
        return {
            "name": name,
            "passed": False,
            "detail": f"{len(empty_tables)} 张表无数据行：{empty_tables}",
        }
    return {"name": name, "passed": True, "detail": "所有产出表均含至少 1 行数据"}


def _check4_jt_consistency(
    hier_rows: list[tuple],
    history: Sequence[dict[str, Any]],
    new_major_rows: list[tuple],
    seed: int = 20260623,
) -> AuditCheck:
    """Random ≥30 matched rows' J/T agree with source (precision-aware).

    Matched rows (匹配阶段 ∈ {严格匹配, 核心名匹配, agent 语义匹配}) are compared against
    the source近三年 history value at (school, major) with a tight tolerance
    (handles float display drift; single-year T=None must match None).

    新增估算 rows (匹配阶段 == 新增专业) are compared against the value in
    ``今年新增往年没有的专业.xlsx`` at (school, major) — the estimate table is the ground
    truth for new-major rows (V5-6 precision split).
    """
    name = "jt_consistency"
    hist_idx = _history_index(history)
    # Index the new-major table by (学校, 专业) → (J, T).
    nm_idx: dict[tuple[str, str], tuple[Any, Any]] = {}
    for r in new_major_rows:
        if len(r) < 5:
            continue
        school = r[0]
        major = r[1]
        j = r[3]
        t = r[4]
        nm_idx[(str(school or ""), str(major or ""))] = (j, t)

    if not hier_rows:
        return {"name": name, "passed": False, "detail": "分层版无数据行"}
    header = tuple(h if h is not None else "" for h in hier_rows[0])
    stage_idx = _stage_col_index(header)

    # Collect candidate matched + estimate rows from the hierarchical output.
    matched_candidates: list[tuple] = []
    estimate_candidates: list[tuple] = []
    for row_idx_0, row in enumerate(hier_rows):
        if row_idx_0 == 0:
            continue
        if not _is_major_row(tuple(row)):
            continue
        stage = _stage_of(tuple(row), stage_idx)
        if stage in MATCHED_STAGES:
            matched_candidates.append(tuple(row))
        elif stage == ESTIMATE_STAGE:
            estimate_candidates.append(tuple(row))

    rng = random.Random(seed)
    sample_matched = rng.sample(
        matched_candidates, min(SAMPLE_SIZE, len(matched_candidates))
    )
    sample_estimate = rng.sample(
        estimate_candidates, min(SAMPLE_SIZE, len(estimate_candidates))
    )

    mismatches: list[str] = []

    def _close(a: Any, b: Any) -> bool:
        if a is None and b is None:
            return True
        if a is None or b is None:
            return False
        try:
            return abs(float(a) - float(b)) <= JT_TOLERANCE
        except (TypeError, ValueError):
            return False

    for row in sample_matched:
        school = str(row[COL_SCHOOL - 1] or "")
        major = str(row[COL_MAJOR_NAME - 1] or "")
        out_j = row[COL_J - 1]
        out_t = row[COL_T - 1]
        hs = hist_idx.get((school, major))
        if not hs:
            # Cannot locate history — skip (not a J/T mismatch per se).
            continue
        # Bug #3: same (school, major) may have multiple 招生类别 rows (普通/公安/
        # 师范) with different J/T — output matches if it hits ANY of them.
        if not any(_close(out_j, h.get("J")) and _close(out_t, h.get("T")) for h in hs):
            mismatches.append(
                f"matched J/T {school}/{major}: {out_j}/{out_t}≠"
                f"{[(h.get('J'), h.get('T')) for h in hs]}"
            )

    for row in sample_estimate:
        school = str(row[COL_SCHOOL - 1] or "")
        major = str(row[COL_MAJOR_NAME - 1] or "")
        out_j = row[COL_J - 1]
        out_t = row[COL_T - 1]
        est = nm_idx.get((school, major))
        if est is None:
            continue
        est_j, est_t = est
        if not _close(out_j, est_j):
            mismatches.append(f"estimate J {school}/{major}: {out_j}≠{est_j}")
            continue
        if not _close(out_t, est_t):
            mismatches.append(f"estimate T {school}/{major}: {out_t}≠{est_t}")

    if mismatches:
        return {
            "name": name,
            "passed": False,
            "detail": (
                f"J/T 不一致 {len(mismatches)} 处（抽样 matched={len(sample_matched)}, "
                f"estimate={len(sample_estimate)}）：{mismatches[:5]}"
            ),
        }
    return {
        "name": name,
        "passed": True,
        "detail": (
            f"抽样 {len(sample_matched)} matched + {len(sample_estimate)} estimate 行 "
            f"J/T 与源值一致（容差 {JT_TOLERANCE}）"
        ),
    }


# ---------------------------------------------------------------------------
# sample artefact (manual / human review; NOT a gate)
# ---------------------------------------------------------------------------


def _write_sample(hier_rows: list[tuple], out_path: Path, seed: int = 20260623) -> int:
    """Write ``audit_sample.xlsx``: random SAMPLE_SIZE major rows for human
    semantic review. Returns the number of sampled rows written."""
    majors: list[tuple] = []
    for i, row in enumerate(hier_rows):
        if i == 0:
            continue
        if _is_major_row(tuple(row)):
            majors.append(tuple(row))
    rng = random.Random(seed)
    sample = rng.sample(majors, min(SAMPLE_SIZE, len(majors)))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "audit_sample"
    if hier_rows:
        ws.append(list(hier_rows[0]))
    for r in sample:
        ws.append(list(r))
    wb.save(out_path)
    wb.close()
    return len(sample)


# 官方来源域名（不认第三方百科/新闻）。run13 fresh-test：agent 第一版「链接待补」
# 照样过审计 → 加这项自动抓「is_rename=true 行缺官方链接」。
_OFFICIAL_LINK_MARKERS = ("moe.gov.cn", "gov.cn", ".edu.cn")


def _check5_rename_official_link(out_dir: Path) -> AuditCheck:
    """改名表每一行都必须含官方来源链接（moe.gov.cn / gov.cn / .edu.cn）。

    改名是 agent 网查判的、note 直填备注列；用户看改名依据只靠这一列。缺链接
    = 网查没做或没留证 → 审计拦下（run13 第一版「链接待补」就漏过了）。
    列名不固定（新旧版表头不同），所以扫**整行所有单元格**找链接，更稳。
    """
    name = "rename_official_link"
    rename_path = out_dir / RENAME_NAME
    if not rename_path.exists():
        return {"name": name, "passed": True, "detail": "无改名表（无改名校）"}

    wb = openpyxl.load_workbook(rename_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) <= 1:  # 仅表头 / 空
        return {"name": name, "passed": True, "detail": "改名表无数据行"}

    missing: list[str] = []
    for row_idx_0, row in enumerate(rows[1:], start=2):  # 数据行，2-based
        cells = [str(c) for c in row if c is not None]
        joined = " ".join(cells)
        if not any(marker in joined for marker in _OFFICIAL_LINK_MARKERS):
            # 报新校名（第 1 列）方便定位
            new_school = cells[0] if cells else f"行{row_idx_0}"
            missing.append(new_school)
    if missing:
        return {
            "name": name,
            "passed": False,
            "detail": f"改名表 {len(missing)} 行缺官方链接(moe/gov/edu)：{missing[:5]}",
        }
    return {
        "name": name,
        "passed": True,
        "detail": f"改名表 {len(rows) - 1} 行均含官方来源链接",
    }


# ---------------------------------------------------------------------------
# audit
# ---------------------------------------------------------------------------


def audit(
    output_dir: str | Path,
    *,
    data_dir: str | Path,
    intermediate_dir: str | Path,
    history: Sequence[dict[str, Any]] | None = None,
    semantic_dir: str | Path | None = None,
) -> AuditReport:
    """Run the six data-quality checks over ``output_dir``.

    Parameters
    ----------
    output_dir
        Directory holding the produced xlsx (hierarchical / flat / edge tables).
    data_dir
        Directory holding the three source xlsx. Used to rebuild the unified
        history when ``history`` is not supplied (check 4 J/T comparison).
    intermediate_dir
        Intermediate-artefact dir (stage files). Currently informational; kept
        in the signature per Plan v2 binding (``能读/重算 history``).
    history
        Optional pre-built unified history (avoids a rebuild). When ``None``
        the history is rebuilt from ``data_dir`` via :func:`build_unified_history`.
    semantic_dir
        Directory holding ``verify_*_result.jsonl`` (default: ``<repo>/semantic-match``).
    """
    out_dir = Path(output_dir)
    data_dir = Path(data_dir)
    sem_dir = (
        Path(semantic_dir)
        if semantic_dir is not None
        else (out_dir.parent / "semantic-match")
    )

    hier_path = out_dir / HIER_ARCHICAL_NAME
    flat_path = out_dir / FLAT_NAME
    hier_rows = _load_rows(hier_path)
    flat_rows = _load_rows(flat_path)

    verdicts = _load_verify_verdicts(sem_dir)

    checks: list[AuditCheck] = []
    checks.append(_check0_judgmental_coverage(hier_rows, verdicts))
    checks.append(_check1_nonempty_stage(flat_rows))
    checks.append(_check2_no_empty_rows(out_dir))
    checks.append(_check3_tables_nonempty(out_dir))

    # History for check 4 (rebuild if not supplied).
    hist = history
    if hist is None:
        try:
            hist = _rebuild_history(data_dir)
        except Exception as exc:  # pragma: no cover — defensive
            logger.warning(
                "audit: history 重建失败 (%s)；check4 将跳过 matched 比对", exc
            )
            hist = []
    new_major_rows: list[tuple] = []
    nm_path = out_dir / NEW_MAJOR_NAME
    if nm_path.exists():
        new_major_rows = _load_rows(nm_path)
    checks.append(_check4_jt_consistency(hier_rows, hist, new_major_rows))
    checks.append(_check5_rename_official_link(out_dir))

    # Human-review sample (does NOT affect ok).
    try:
        _write_sample(hier_rows, out_dir / "audit_sample.xlsx")
    except Exception as exc:  # pragma: no cover — defensive
        logger.warning("audit: audit_sample.xlsx 写出失败 (%s)", exc)

    ok = all(c["passed"] for c in checks)
    return AuditReport(ok=ok, checks=checks)


def _rebuild_history(data_dir: Path) -> list[dict[str, Any]]:
    """Rebuild the unified history from the three source workbooks.

    Mirrors :func:`scripts.run_pipeline._load_rows` + Stage 0. Kept local so
    the audit does not depend on the pipeline runner (which has side effects).
    """
    from scripts.constants import J3_SHEET
    from scripts.run_pipeline import SOURCE_FILES

    def _load(path: Path, sheet: str | None = None):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        try:
            return list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

    j3_rows = _load(data_dir / SOURCE_FILES["j3"], J3_SHEET)
    tq_rows = _load(data_dir / SOURCE_FILES["tq"])
    return build_unified_history(j3_rows, tq_rows)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    """CLI entry: ``python -m scripts.audit_output``.

    Exits 0 iff every check passes (``ok=True``), non-zero otherwise.
    """
    parser = argparse.ArgumentParser(
        prog="python -m scripts.audit_output",
        description="V5-3 data-quality audit hard gate over real produced xlsx.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output"),
        help="Directory holding the produced xlsx (default: output)",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=Path("data"),
        help="Directory holding the three source xlsx (default: data)",
    )
    parser.add_argument(
        "--intermediate-dir",
        type=Path,
        default=Path("intermediate"),
        help="Intermediate-artefact directory (default: intermediate)",
    )
    parser.add_argument(
        "--semantic-dir",
        type=Path,
        default=Path("semantic-match"),
        help="Directory holding verify_*_result.jsonl (default: semantic-match)",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    report = audit(
        args.output_dir,
        data_dir=args.data_dir,
        intermediate_dir=args.intermediate_dir,
        semantic_dir=args.semantic_dir,
    )
    print("=== 数据质量审计 ===")
    for c in report.checks:
        flag = "PASS" if c["passed"] else "FAIL"
        print(f"[{flag}] {c['name']}: {c['detail']}")
    print(f"=== 结果: {'OK (exit 0)' if report.ok else 'FAIL (exit 1)'} ===")
    raise SystemExit(0 if report.ok else 1)


if __name__ == "__main__":
    main()
