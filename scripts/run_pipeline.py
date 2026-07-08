"""Slice 7 — end-to-end pipeline runner.

Threads the **deterministic** stages of the admission-data pipeline (spec §6,
Plan v2 Slice 7 / issue #8):

    Stage 0  build_unified_history + build_dagluben (常规批 + 提前批, 专科 excluded)
       ↓
    Stage 1  match_strict                       (严格 3-tuple)
       ↓
    Stage 2  (agent — harness-side)
       - if semantic-match/batch_*_result.jsonl exist → apply_results
       - else → emit batch_NN_prompt.json + log「Stage2 待 harness 派发」
       ↓
    Stage 3  identify_new_majors + estimate + write_new_major_table
       ↓
    Stage 3  edges:
       - rename: if semantic-match/rename_result.jsonl exists → apply_rename
                 else → empty renamed set + log「改名 待 harness 派发」
       - deleted_majors (uses the confirmed renamed set to exclude)
       - flight_and_special (FLIGHT_BATCH unmatched + remaining unmatched)
       ↓
    write_outputs  (hierarchical + flat — same MatchResult source)
    write_edge_tables (被删/新增校/停招消失校/特殊/改名)

The agent (Stage 2 semantic, Stage 3 rename pairing) and WebSearch (rename
remarks) are **harness-side steps** — Python cannot invoke the Agent/WebSearch
tools. This module therefore treats them as optional: when their result jsonls
are absent, the run still succeeds end-to-end and emits the prompts + candidates
the harness needs to dispatch them.

Pure function: :func:`run` returns a structured report dict and writes all
artefacts under ``out_dir`` / ``semantic_dir`` / ``intermediate_dir``. The three
source files are guarded by :func:`io_source.assert_unchanged` before and after.
"""

from __future__ import annotations

import argparse
import collections
import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scripts import io_source
from scripts.constants import FLIGHT_BATCH, J3_SHEET
from scripts.models import DaglubenRow, EstimateResult, HistoryRow, MatchResult
from scripts.rename_detect import (
    apply_rename,
    prep_rename_candidates,
    write_rename_prompt,
)
from scripts.stage0_merge import (
    build_dagluben_early,
    build_dagluben_regular,
    build_unified_history,
)
from scripts.stage1_strict import match_strict
from scripts.stage2_agent import build_batches, write_prompts
from scripts.stage2_apply import apply_results
from scripts.stage3_edges import deleted_majors, flight_and_special
from scripts.stage3_newmajor import estimate
from scripts.write_edge_tables import (
    identify_new_majors,
    write_deleted_major_table,
    write_gone_school_table,
    write_new_major_table,
    write_new_school_table,
    write_rename_table,
    write_special_table,
)
from scripts.write_outputs import write_flat, write_hierarchical

__all__ = ["run", "main", "SOURCE_FILES", "PipelineReport"]

logger = logging.getLogger(__name__)

# Repository layout — the source filenames (spec §2 / CLAUDE.md).
SOURCE_FILES: dict[str, str] = {
    "j3": "近三年学校批次专业线差统计.xlsx",
    "tq": "山东省高考提前批录取数据.xlsx",
    "dl": "山东省2026年大绿本招生计划.xlsx",
}
# 可选源——不一定存在。提前批补充表（tq）只是补 J3 没统计到的提前批专业；
# J3 已统计的提前批用现成线差。补充表不存在时跳过，不报错（不特化「必须有」）。
SOURCE_OPTIONAL: frozenset[str] = frozenset({"tq"})

# Typed alias for the structured report returned by :func:`run`. Kept as a
# plain dict (not TypedDict) because the report is a test/debug surface, not a
# stable inter-module contract.
PipelineReport = dict[str, Any]


def _load_rows(path: Path, sheet_name: str | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    try:
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _guard_sources(
    data_dir: Path, source_files: dict[str, str] | None = None
) -> dict[str, str]:
    """Hash every source before the run; return the baseline map.

    可选源（SOURCE_OPTIONAL）不存在时跳过，不强求（如提前批补充表）。
    source_files 覆盖 SOURCE_FILES（CLI 参数化文件名）。
    """
    files = source_files or SOURCE_FILES
    hashes: dict[str, str] = {}
    for key, name in files.items():
        p = data_dir / name
        if key in SOURCE_OPTIONAL and not p.exists():
            continue
        h = io_source.sha256(p)
        hashes[name] = h
        io_source.assert_unchanged(p, h)
    return hashes


def _apply_stage2(
    post_coarse_unmatched: list[DaglubenRow],
    history: list[HistoryRow],
    semantic_dir: Path,
    with_agent_results: bool,
) -> tuple[list[MatchResult], list[DaglubenRow], bool]:
    """Stage 2 — apply agent jsonl if present, else emit prompts + log.

    Returns ``(semantic_results, still_unmatched, applied)``:
      - ``semantic_results``: MatchResult rows produced by the agent (empty if
        no jsonl was applied).
      - ``still_unmatched``: dagluben rows the agent did NOT resolve (input to
        Stage 3 new-major / edges). When no jsonl is applied this equals the
        full ``post_coarse_unmatched`` list.
      - ``applied``: True iff at least one result line was read and applied.
    """
    semantic_dir.mkdir(parents=True, exist_ok=True)

    if not post_coarse_unmatched:
        # Nothing for the agent to do; still emit no prompts, no apply.
        logger.info("Stage2: Stage 1.5 后无剩余未匹配行，无需 agent 派发")
        return [], [], False

    result_paths = sorted(semantic_dir.glob("batch_*_result.jsonl"))
    if with_agent_results and result_paths:
        applied_results = apply_results(result_paths, post_coarse_unmatched, history)
        resolved_idx = {r["src_row_idx"] for r in applied_results if r.get("matched")}
        still = [
            d
            for d in post_coarse_unmatched
            if d["src_row_idx"] not in resolved_idx
            and not any(
                r["src_row_idx"] == d["src_row_idx"] and r.get("matched")
                for r in applied_results
            )
        ]
        logger.info(
            "Stage2: 已应用 %d 条 agent 结果（%d 条命中，%d 条仍未匹配）",
            len(applied_results),
            len(resolved_idx),
            len(still),
        )
        return applied_results, still, True

    # No results to apply → emit batch prompts for harness dispatch.
    # ISSUE-5: 先清掉旧的 batch_*_prompt.json——改名/数据变化后批次数会变，
    # 残留的 stale prompt 带过时候选，被 agent 误派会用错数据（run_stage_verify_prep
    # 已对 verify_batch_*.json 这么做，照搬）。
    for stale in sorted(semantic_dir.glob("batch_*_prompt.json")):
        stale.unlink()
    batches = build_batches(post_coarse_unmatched, history, batch_size=20)
    write_prompts(batches, semantic_dir)
    logger.info(
        "Stage2: 未发现 batch_*_result.jsonl，已写出 %d 个批次 prompt；"
        "Stage2 待 harness 派发（见 semantic-match/RUN.md）",
        len(batches),
    )
    return [], list(post_coarse_unmatched), False


def _apply_rename(
    dagluben: list[DaglubenRow],
    history: list[HistoryRow],
    semantic_dir: Path,
    with_agent_results: bool,
) -> tuple[list, set[str], bool]:
    """Rename agent step — apply rename_result.jsonl if present.

    Returns ``(rename_rows, renamed_dgl_schools, applied)``. When no jsonl is
    present, emits rename candidates + prompt and returns an empty renamed set
    (so被删 uses the upper bound — true被删 ⊆ the reported count after the
    agent excludes改名校).
    """
    semantic_dir.mkdir(parents=True, exist_ok=True)
    dgl_schools = sorted({d["school"] for d in dagluben if d.get("school")})
    hist_schools = sorted({h["school"] for h in history if h.get("school")})

    candidates = prep_rename_candidates(dgl_schools, hist_schools, topk=0)
    write_rename_prompt(candidates, semantic_dir)

    rename_path = semantic_dir / "rename_result.jsonl"
    if with_agent_results and rename_path.exists():
        rename_rows, confirmed = apply_rename([rename_path], dagluben, history)
        logger.info("改名: 已应用 rename_result.jsonl（%d 所改名校）", len(confirmed))
        return rename_rows, confirmed, True

    logger.info(
        "改名: 未发现 rename_result.jsonl（候选 %d 所）；改名 待 harness 派发"
        "（见 research/RUN_RENAME.md）",
        len(candidates),
    )
    return [], set(), False


def _extract_official_url(text: str) -> str:
    """第一个官方来源 URL（教育部 moe.gov.cn / 省政府 .gov.cn）。不认第三方百科/新闻。"""
    import re

    for line in text.splitlines():
        for m in re.findall(r"https?://[^\s)）]+", line):
            if "moe.gov.cn" in m or ".gov.cn" in m:
                return m
    return ""


def _research_summary(text: str, school: str) -> str:
    """research md →「<结论> 来源：<官方 URL>」。

    结论取含 前身/更名/原名/转设/合并 等的关键行；来源只认官方（moe.gov.cn /
    省政府公文 .gov.cn），**不认**第三方百科/新闻。无官方来源时只给结论。
    """
    concl_kws = ("前身", "更名", "原名", "转设", "合并", "升格", "揭牌", "教育部", "由")
    skip_kws = ("查询语句", "查询")
    conclusion = ""
    cands = []
    for line in text.splitlines():
        s = line.strip().lstrip("-").lstrip("#").strip()
        if not s or any(k in s[:8] for k in skip_kws):
            continue
        if any(k in s for k in concl_kws):
            cands.append(s)
    if cands:
        conclusion = max(cands, key=len)[:100]
    url = _extract_official_url(text)
    parts = [p for p in (conclusion, f"来源：{url}" if url else "") if p]
    return " ".join(parts) or f"（见 research/{school}.md）"


def _enrich_rename_rows(
    rename_rows: list,
    dagluben: list,
    research_dir: str = "research",
) -> None:
    """Populate ``major_count_2026`` + websearch ``remark`` on rename rows
    (spec §6 Task 6.3) before the学校改名表 is written. apply_rename returns
    RenameRows without these fields.

    #7: research md 可能用新校名或旧校名命名，都试；备注格式「结论 + 官方来源
    URL」；merge_remark 保护 manual_reviewed（人工核验过的不被网查覆盖）。
    """
    from scripts.rename_websearch import merge_remark

    cnt = collections.Counter(d.get("school", "") for d in dagluben if d.get("school"))
    rdir = Path(research_dir)
    for r in rename_rows:
        ns = r.get("new_school", "")
        os_ = r.get("old_school", "")
        r["major_count_2026"] = cnt.get(ns, 0)
        # BUG-4: agent 在 rename_result.jsonl 填的 note（结论+官方链接）已进 remark，
        # 是改名表备注的权威来源——有就不覆盖。没有才回退到 research/<校名>.md 扫描。
        if (r.get("remark") or "").strip():
            continue
        # research md 可能用新校名或旧校名命名，逐一试，取第一个有实质内容的。
        summary = ""
        for name in (ns, os_):
            if not name:
                continue
            md = rdir / f"{name}.md"
            if md.exists():
                s = _research_summary(md.read_text(encoding="utf-8"), name)
                if s and not s.startswith("（见"):
                    summary = s
                    break
        if summary:
            # merge_remark 尊重 manual_reviewed（人工核验过的备注不被覆盖）。
            r["remark"] = merge_remark(summary, r)["remark"]
        else:
            r.setdefault("remark", f"（见 research/{ns}.md）")


def _build_main_results(
    dagluben: list[DaglubenRow],
    strict_results: list[MatchResult],
    coarse_results: list[MatchResult],
    semantic_results: list[MatchResult],
    new_major_estimates: dict[int, EstimateResult],
    *,
    classified_idx: set[int] | None = None,
) -> list[MatchResult]:
    """Assemble the final per-row MatchResult list, one entry per大绿本 row.

    Resolution order (first matched wins per src_row_idx): strict → coarse →
    semantic. Unmatched results never claim a slot. Rows that survived all
    matchers pick up the new-major estimate (if any), otherwise the special
    fallback (#6c: rename-pending marker removed — renamed schools now match
    via renamed history in Stage 1-2, no J/T-empty pending rows).

    ``classified_idx`` (Plan v2 阻断2) optionally overrides the classified
    set — the V5-0 demote step shrinks it (drops存疑 idx) BEFORE this build so
    demoted rows fall naturally into the unmatched / special bucket. When
    ``None`` it is derived from the matched results as usual.
    """
    by_idx: dict[int, MatchResult] = {}

    # When the V5-0 demote step passes an explicit classified set, only
    # classified idx may claim a matched slot (存疑 idx were stripped upstream
    # and must fall through to the unmatched bucket).
    allowed = classified_idx

    for r in strict_results:
        if r.get("matched") and (allowed is None or r["src_row_idx"] in allowed):
            by_idx.setdefault(r["src_row_idx"], r)
    for r in coarse_results:
        if r.get("matched") and (allowed is None or r["src_row_idx"] in allowed):
            by_idx.setdefault(r["src_row_idx"], r)
    for r in semantic_results:
        if r.get("matched") and (allowed is None or r["src_row_idx"] in allowed):
            by_idx.setdefault(r["src_row_idx"], r)

    # Remaining rows: new-major estimate or rename-pending marker.
    for d in dagluben:
        idx = d["src_row_idx"]
        if idx in by_idx:
            continue
        school = d.get("school", "")
        est = new_major_estimates.get(idx)
        if est is not None:
            by_idx[idx] = MatchResult(
                src_row_idx=idx,
                school=school,
                school_cat=d.get("school_cat", ""),
                major=d.get("major", ""),
                matched=False,
                J=est.get("value"),
                T=est.get("T"),
                log=est.get("log", ""),
            )
        else:
            # 特殊: 未匹配本科。spec 要求每行都有日志，兜底发特殊情况日志
            # （详情见 output/未能匹配的专业.xlsx）。
            from scripts.constants import LOG_SPECIAL_UNMATCHED

            by_idx[idx] = MatchResult(
                src_row_idx=idx,
                school=school,
                school_cat=d.get("school_cat", ""),
                major=d.get("major", ""),
                matched=False,
                J=None,
                T=None,
                log=LOG_SPECIAL_UNMATCHED,
            )

    # Emit in dagluben order for stable downstream output. Every 本科 row now
    # carries a MatchResult (matched / 新增 / 改名 / 特殊), so no row is left
    # without a 匹配日志.
    return [by_idx[d["src_row_idx"]] for d in dagluben]


_ONE_LINE_PAIR_RE = re.compile(r"(20\d{2})\s*=\s*(\d{2,3})")


def read_one_line_from_notes(
    j3_path: str | Path, notes_sheet: str = "说明"
) -> dict[int, int] | None:
    """从近三年统计表「说明」sheet 自动读一段线。

    找到文本含「一段线」的单元格，抽其中的「年份=分数」对（≥2 年才算数）。
    找不到返回 None，调用方退到 ``constants.ONE_LINE`` 默认。根除 BUG-1 那类
    硬编码——数据年年变，一段线写在说明 sheet 里就该自动用，不靠手改 help/常量。
    """
    path = Path(j3_path)
    if not path.exists():
        return None
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = [notes_sheet] if notes_sheet in wb.sheetnames else list(wb.sheetnames)
        for sn in sheets:
            for row in wb[sn].iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and "一段线" in cell:
                        pairs = {
                            int(y): int(v) for y, v in _ONE_LINE_PAIR_RE.findall(cell)
                        }
                        if len(pairs) >= 2:
                            return pairs
    finally:
        wb.close()
    return None


def run(
    data_dir: str | Path,
    out_dir: str | Path,
    *,
    with_agent_results: bool = False,
    semantic_dir: str | Path | None = None,
    intermediate_dir: str | Path | None = None,
    source_files: dict[str, str] | None = None,
    one_line: dict[int, int] | None = None,
    supplement_batches: frozenset[str] | None = None,
    supplement_low_cols: dict[int, int] | None = None,
    dagluben_early_batches: frozenset[str] | None = None,
    flight_batch: str | None = None,
) -> PipelineReport:
    """Run the deterministic admission-data pipeline end-to-end.

    Parameters
    ----------
    data_dir
        Directory holding the three source xlsx (see :data:`SOURCE_FILES`).
    out_dir
        Directory where the final xlsx outputs are written.
    with_agent_results
        When True, apply any ``semantic-match/batch_*_result.jsonl`` and
        ``semantic-match/rename_result.jsonl`` produced by the harness agent
        step. When False (default), the run emits prompts + candidates and
        logs that the agent step is pending.
    semantic_dir
        Where to read/write agent prompts/results. Defaults to ``<repo>/semantic-match``.
    intermediate_dir
        Where to write intermediate CSVs. Defaults to ``<repo>/intermediate``.

    Returns
    -------
    PipelineReport
        Structured dict with source hashes, dagluben row indices, main
        MatchResult list, edge buckets, new-major rows, coverage stats, and
        flags (stage2_applied / renamed_dgl_schools). See the field assignments
        below for the full shape.
    """
    data_dir = Path(data_dir)
    out_dir = Path(out_dir)
    if semantic_dir is None:
        semantic_dir = data_dir.parent / "semantic-match"
    else:
        semantic_dir = Path(semantic_dir)
    if intermediate_dir is None:
        intermediate_dir = data_dir.parent / "intermediate"
    else:
        intermediate_dir = Path(intermediate_dir)

    out_dir.mkdir(parents=True, exist_ok=True)
    intermediate_dir.mkdir(parents=True, exist_ok=True)

    # --- Source immutability guard (before) ---------------------------------
    files = source_files or SOURCE_FILES
    source_hashes = _guard_sources(data_dir, files)

    # --- Stage 0 ------------------------------------------------------------
    j3_rows = _load_rows(data_dir / files["j3"], J3_SHEET)
    tq_path = data_dir / files["tq"]
    tq_rows = _load_rows(tq_path) if tq_path.exists() else []
    dl_rows = _load_rows(data_dir / files["dl"])

    # 未显式传 one_line 时，自动从「说明」sheet 读一段线——数据年年变，写在
    # 说明 sheet 里就该自动用，不靠手改 help/常量（BUG-1 那类 Bug 的根除）。
    if one_line is None:
        auto_one_line = read_one_line_from_notes(data_dir / files["j3"])
        if auto_one_line:
            one_line = auto_one_line
            logger.info("一段线: 自动从「说明」sheet 读取 %s", auto_one_line)

    history = build_unified_history(
        j3_rows,
        tq_rows,
        one_line=one_line,
        batches=supplement_batches,
        low_cols=supplement_low_cols,
    )
    dagluben = [
        *build_dagluben_regular(dl_rows),
        *build_dagluben_early(dl_rows, batches=dagluben_early_batches),
    ]
    dgl_indices = [d["src_row_idx"] for d in dagluben]
    logger.info(
        "Stage0: 统一历史 %d 行；大绿本本科专业 %d 行",
        len(history),
        len(dagluben),
    )

    # --- 改名检测（#6c 提前到 Stage 0 后）------------------------------------
    # rename apply 必须在 Stage 1 前：把旧校名的 history rows 的 school 替换为
    # 新校名 → Stage 1-2 自然匹配改名校专业的旧名线差（不再 J/T 留空 pending）。
    rename_rows, renamed_dgl_schools, rename_applied = _apply_rename(
        dagluben,
        history,
        semantic_dir,
        with_agent_results,
    )
    if renamed_dgl_schools:
        old_to_new = {
            r.get("old_school", ""): r.get("new_school", "")
            for r in rename_rows
            if r.get("old_school")
        }
        history = [
            {**h, "school": old_to_new.get(h.get("school", ""), h.get("school", ""))}
            for h in history
        ]
        logger.info(
            "改名联动: %d 所旧校名 history 并入新校名（Stage 1-2 自动用旧名数据）",
            len(old_to_new),
        )

    # --- Stage 1 (strict) ---------------------------------------------------
    strict_results = match_strict(dagluben, history)
    strict_unmatched_dgl = [
        d for d, r in zip(dagluben, strict_results, strict=True) if not r.get("matched")
    ]
    logger.info(
        "Stage1: 严格匹配 %d/%d (%.1f%%)",
        sum(1 for r in strict_results if r.get("matched")),
        len(strict_results),
        100.0
        * sum(1 for r in strict_results if r.get("matched"))
        / max(1, len(strict_results)),
    )

    # --- Stage 1.5 已并入 Stage 2 ------------------------------------------
    # 程序模糊匹配（核心名粗筛自动接受）已停用：全部 strict 未匹配交由 Stage 2
    # agent 逐条语义判断。skill 规格＝只有「程序严格匹配 + agent 语义匹配」，
    # 不让程序做模糊决定（避免「临床医学5+3」式低级误判）。coarse_results 恒空，
    # 保留为占位以维持下游 _build_main_results 签名稳定（#18 清理时再删）。
    coarse_results: list[MatchResult] = []
    post_coarse_unmatched = strict_unmatched_dgl
    logger.info(
        "Stage1.5: 程序模糊匹配已停用，%d 条 strict 未匹配全部进 Stage2 agent",
        len(post_coarse_unmatched),
    )

    # --- Stage 2 (agent) ----------------------------------------------------
    semantic_results, post_stage2_unmatched, stage2_applied = _apply_stage2(
        post_coarse_unmatched,
        history,
        semantic_dir,
        with_agent_results,
    )

    # --- Stage 3 (new-major estimation) ------------------------------------
    new_majors = identify_new_majors(post_stage2_unmatched, history)
    by_school: dict[str, list[HistoryRow]] = collections.defaultdict(list)
    for h in history:
        by_school[h["school"]].append(h)

    new_major_estimates: dict[int, EstimateResult] = {}
    new_major_rows: list[dict[str, Any]] = []
    for d in new_majors:
        est = estimate(d, by_school.get(d["school"], []))
        new_major_estimates[d["src_row_idx"]] = est
        new_major_rows.append(
            {
                "src_row_idx": d["src_row_idx"],
                "school": d.get("school", ""),
                "major": d.get("major", ""),
                "subject": d.get("subject", ""),
                "value": est.get("value"),
                "T": est.get("T"),
                "level": est.get("level"),
                "n": est.get("n"),
                "log": est.get("log", ""),
            }
        )
    write_new_major_table(new_major_rows, out_dir / "今年新增往年没有的专业.xlsx")
    logger.info("Stage3 新增专业: %d", len(new_majors))

    # --- 改名表写出（_apply_rename 已在 Stage 0 后跑；这里只 enrich + write）---
    _enrich_rename_rows(rename_rows, dagluben)
    write_rename_table(rename_rows, out_dir / "学校改名表.xlsx")

    # --- Stage 3 (edges: deleted / flight / special) -----------------------
    dgl_present = {d["school"] for d in dagluben if d.get("school")}
    dgl_school_major = {(d.get("school", ""), d.get("major", "")) for d in dagluben}
    deleted_pool = deleted_majors(history, dgl_present, renamed_dgl_schools)
    true_deleted = [
        dm
        for dm in deleted_pool
        if (dm.get("school", ""), dm.get("major", "")) not in dgl_school_major
    ]
    write_deleted_major_table(true_deleted, out_dir / "往年有但今年停招的专业.xlsx")

    # New-school / gone-school tables: 大绿本独有校 / 历史独有校 minus rename.
    major_count: dict[str, int] = collections.Counter(
        d.get("school", "") for d in dagluben if d.get("school")
    )
    hist_school_set = {h["school"] for h in history if h.get("school")}
    dgl_unique = [
        s for s in sorted(dgl_present - hist_school_set) if s not in renamed_dgl_schools
    ]
    # 改名旧校名从停招消失校表移出（它们是改名的旧名，不是真停招）。
    rename_old_schools = {
        r.get("old_school", "") for r in rename_rows if r.get("old_school")
    }
    hist_unique = [
        s for s in sorted(hist_school_set - dgl_present) if s not in rename_old_schools
    ]
    write_new_school_table(
        [{"new_school": s, "major_count_2026": major_count[s]} for s in dgl_unique],
        out_dir / "今年新招生的学校.xlsx",
    )
    write_gone_school_table(
        [{"old_school": s} for s in hist_unique],
        out_dir / "往年有今年停招的学校.xlsx",
    )

    # --- V5-0 second-pass verification apply (Plan v2 阻断2) ----------------
    # If verify_*_result.jsonl exists, apply verdicts BEFORE _build_main_results:
    # filter 存疑 idx out of coarse/semantic results + classified_idx so they
    # fall naturally into remaining_unmatched → special, carrying a
    # 「复核存疑：<原因>」 log (bypassing the generic fallback).
    verify_result_paths = sorted(semantic_dir.glob("verify_*_result.jsonl"))
    verdict_by_idx: dict[int, str] = {}
    demoted_map: dict[int, str] = {}
    coarse_for_main = coarse_results
    semantic_for_main = semantic_results
    classified_for_main: set[int] | None = None
    verify_applied = False
    if with_agent_results and verify_result_paths:
        from scripts.verify_judgment import apply_verify, filter_demoted

        judgmental = list(
            semantic_results
        )  # coarse 已停用，判断型匹配只剩 agent 语义匹配
        verify_out = apply_verify(verify_result_paths, dagluben, judgmental)
        verdict_by_idx = verify_out["verdict_by_idx"]
        # reasons map: pull the reason from the demoted EdgeRows (built by apply_verify).
        reasons = {
            e["src_row_idx"]: e["log"].split("：", 1)[1]
            if "：" in e.get("log", "")
            else ""
            for e in verify_out["demoted"]
        }
        # coarse 已停用（coarse_results 恒空）→ 无需 filter；coarse_for_main 保持空。
        coarse_for_main: list[MatchResult] = []
        semantic_for_main, _, _ = filter_demoted(
            semantic_results,
            set(),
            verdict_by_idx,
            reasons,
        )
        # Build the classified set from the filtered results (so存疑 idx drop).
        classified_for_main = (
            {r["src_row_idx"] for r in strict_results if r.get("matched")}
            | {r["src_row_idx"] for r in semantic_for_main if r.get("matched")}
            | {d["src_row_idx"] for d in new_majors}
        )
        demoted_map = {
            idx: reasons.get(idx, "")
            for idx, v in verdict_by_idx.items()
            if v == "存疑"
        }
        verify_applied = True
        logger.info(
            "复核: 已应用 %d 条复核结果（%d 确定，%d 存疑→特殊）",
            len(verdict_by_idx),
            sum(1 for v in verdict_by_idx.values() if v == "确定"),
            len(demoted_map),
        )
    else:
        if verify_result_paths:
            logger.info(
                "复核: 检测到 verify_*_result.jsonl 但未启用 --with-agent-results，跳过 apply"
            )
        else:
            logger.info(
                "复核: 未发现 verify_*_result.jsonl，判断型复核 待 harness 派发"
                "（见 semantic-match/RUN_VERIFY.md）"
            )

    # Flight + remaining-unmatched → special.
    classified_idx = (
        classified_for_main
        if classified_for_main is not None
        else (
            {r["src_row_idx"] for r in strict_results if r.get("matched")}
            | {r["src_row_idx"] for r in semantic_results if r.get("matched")}
            | {d["src_row_idx"] for d in new_majors}
        )
    )
    remaining_unmatched = [
        d for d in dagluben if d["src_row_idx"] not in classified_idx
    ]
    _flight_label = flight_batch or FLIGHT_BATCH
    flight = [d for d in remaining_unmatched if d.get("batch") == _flight_label]
    other = [d for d in remaining_unmatched if d.get("batch") != _flight_label]
    # 「对不上」的 other 行（同核心多对一/类别冲突/大类无有效对应）→ 按同校同选科
    # 均值估算（用户口径 2026-07-08：不留在表里空着）。飞行/无历史的不在此列。
    other_estimates: dict[int, EstimateResult] = {
        d["src_row_idx"]: estimate(d, by_school.get(d.get("school", ""), []))
        for d in other
    }
    special_rows = flight_and_special(
        flight,
        other,
        demoted_map=demoted_map,
        history=history,
        estimates=other_estimates,
    )
    write_special_table(special_rows, out_dir / "未能匹配的专业.xlsx")

    # --- Outputs (hierarchical + flat, same MatchResult source) ------------
    main_results = _build_main_results(
        dagluben,
        strict_results,
        coarse_for_main,
        semantic_for_main,
        new_major_estimates,
        classified_idx=classified_for_main,
    )
    dl_path = data_dir / files["dl"]
    write_hierarchical(
        dl_path,
        main_results,
        out_dir / "大绿本_完整版_含线差.xlsx",
    )
    write_flat(
        dl_path,
        main_results,
        out_dir / "大绿本_专业列表_含线差.xlsx",
    )

    # --- Source immutability guard (after) ---------------------------------
    for name, h in source_hashes.items():
        io_source.assert_unchanged(data_dir / name, h)

    # --- Coverage report ---------------------------------------------------
    matched_n = sum(1 for r in main_results if r.get("matched"))
    coverage = {
        "total_dagluben": len(dagluben),
        "matched": matched_n,
        "new_major": len(new_majors),
        "special": len(special_rows),
        "deleted": len(true_deleted),
        "new_school": len(dgl_unique),
        "gone_school": len(hist_unique),
        "rename": len(renamed_dgl_schools),
    }

    logger.info(
        "覆盖率: 匹配 %d / 新增 %d / 特殊 %d / 被删 %d / 改名 %d （总 %d）",
        coverage["matched"],
        coverage["new_major"],
        coverage["special"],
        coverage["deleted"],
        coverage["rename"],
        coverage["total_dagluben"],
    )

    return {
        "source_hashes": source_hashes,
        "dagluben_indices": dgl_indices,
        "dagluben_rows": dagluben,
        "main_results": main_results,
        "history": history,
        "new_major_rows": new_major_rows,
        "post_coarse_unmatched_indices": [
            d["src_row_idx"] for d in post_coarse_unmatched
        ],
        "edge": {
            "deleted": true_deleted,
            "special": special_rows,
            "new_school": dgl_unique,
            "gone_school": hist_unique,
            "rename": rename_rows,
        },
        "coverage": coverage,
        "stage2_applied": stage2_applied,
        "rename_applied": rename_applied,
        "verify_applied": verify_applied,
        "verdict_by_idx": verdict_by_idx,
        "demoted_map": demoted_map,
        "renamed_dgl_schools": renamed_dgl_schools,
    }


def add_source_files_args(parser: argparse.ArgumentParser) -> None:
    """注册数据源 CLI 参数（文件名 / 一段线 / 补充表 / 提前批）。

    ``run_pipeline.main`` 与 ``run_stage_verify_prep.main`` 共用这一组参数 +
    :func:`parse_source_files_args`，单一真理源——避免两入口 drift（BUG-2:
    verify_prep 曾不转发任何 source_files 参数，导致提前批线差用错默认）。
    """
    parser.add_argument("--dl-file", default=None, help="大绿本文件名（覆盖默认）")
    parser.add_argument(
        "--j3-file", default=None, help="近三年统计表文件名（覆盖默认）"
    )
    parser.add_argument(
        "--tq-file",
        default=None,
        help="提前批补充表文件名（覆盖默认；不存在则跳过）",
    )
    parser.add_argument(
        "--one-line",
        default=None,
        help="一段线，格式「年份=分数」逗号分隔，如「2023=443,2024=444,2025=441」"
        "（覆盖默认；不传则自动从近三年表「说明」sheet 读，再退到内置默认）",
    )
    parser.add_argument(
        "--supplement-batches",
        default=None,
        help="补充表批次名，逗号分隔（如 本科提前批A类,本科提前批B类）",
    )
    parser.add_argument(
        "--supplement-low-cols",
        default=None,
        help="补充表低分列，格式「2025=10,2024=14,2023=18」（0 开始数）",
    )
    parser.add_argument(
        "--dagluben-early-batches",
        default=None,
        help="大绿本提前批的批次名，逗号分隔（默认 1.提前批A类,2.提前批B类,3.提前批—飞行技术(军队)）",
    )
    parser.add_argument(
        "--flight-batch",
        default=None,
        help="飞行技术批次名（默认 3.提前批—飞行技术(军队)）",
    )


def parse_source_files_args(args: argparse.Namespace) -> dict[str, object]:
    """把 :func:`add_source_files_args` 注册的参数解析成 ``run()`` 的关键字参数。

    Returns a dict suitable for ``run(..., **<this>)``: source_files / one_line
    / supplement_batches / supplement_low_cols / dagluben_early_batches /
    flight_batch. ``None`` values fall through to ``run()``'s defaults.
    """
    source_files: dict[str, str] | None = None
    if args.dl_file or args.j3_file or args.tq_file:
        source_files = dict(SOURCE_FILES)
        if args.dl_file:
            source_files["dl"] = args.dl_file
        if args.j3_file:
            source_files["j3"] = args.j3_file
        if args.tq_file:
            source_files["tq"] = args.tq_file
    one_line: dict[int, int] | None = None
    if args.one_line:
        one_line = {}
        for pair in args.one_line.split(","):
            year, val = pair.split("=")
            one_line[int(year)] = int(val)
    supplement_batches = (
        frozenset(args.supplement_batches.split(","))
        if args.supplement_batches
        else None
    )
    supplement_low_cols: dict[int, int] | None = None
    if args.supplement_low_cols:
        supplement_low_cols = {}
        for pair in args.supplement_low_cols.split(","):
            year, col = pair.split("=")
            supplement_low_cols[int(year)] = int(col)
    dagluben_early_batches = (
        frozenset(args.dagluben_early_batches.split(","))
        if args.dagluben_early_batches
        else None
    )
    return {
        "source_files": source_files,
        "one_line": one_line,
        "supplement_batches": supplement_batches,
        "supplement_low_cols": supplement_low_cols,
        "dagluben_early_batches": dagluben_early_batches,
        "flight_batch": args.flight_batch,
    }


def main() -> None:
    """CLI entry: run the deterministic chain over the real data dir."""
    parser = argparse.ArgumentParser(
        prog="python -m scripts.run_pipeline",
        description="Run the deterministic admission-data pipeline end-to-end.",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=Path("data"),
        help="Directory holding the three source xlsx (default: data)",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("output"),
        help="Directory for final xlsx outputs (default: output)",
    )
    parser.add_argument(
        "--semantic-dir",
        type=Path,
        default=Path("semantic-match"),
        help="Directory for agent prompts/results (default: semantic-match)",
    )
    parser.add_argument(
        "--intermediate-dir",
        type=Path,
        default=Path("intermediate"),
        help="Directory for intermediate CSVs (default: intermediate)",
    )
    parser.add_argument(
        "--with-agent-results",
        action="store_true",
        help="Apply any batch_*_result.jsonl / rename_result.jsonl present in "
        "semantic-dir. Without this flag, prompts are emitted and the "
        "agent step is logged as pending harness dispatch.",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
    )
    add_source_files_args(parser)
    args = parser.parse_args()

    # CLI 参数化：文件名 + 一段线（覆盖模块默认，agent 传参不手写代码）。
    # 解析逻辑与 run_stage_verify_prep 共用 add_source_files_args /
    # parse_source_files_args，避免两入口 drift（BUG-2: verify_prep 曾不转发）。
    _source_kwargs = parse_source_files_args(args)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    report = run(
        args.data_dir,
        args.out_dir,
        with_agent_results=args.with_agent_results,
        semantic_dir=args.semantic_dir,
        intermediate_dir=args.intermediate_dir,
        **_source_kwargs,
    )

    cov = report["coverage"]
    semantic_suffix = "+语义" if report["stage2_applied"] else ""
    print("=== 管线确定性链完成 ===")
    print(f"大绿本本科专业总数: {cov['total_dagluben']}")
    print(f"  严格匹配{semantic_suffix}: {cov['matched']}")
    print(f"  新增专业(估算):     {cov['new_major']}")
    print(f"  特殊情况:           {cov['special']}")
    print(f"  被删旧专业:         {cov['deleted']}")
    print(f"  新增校:             {cov['new_school']}")
    print(f"  停招消失校:         {cov['gone_school']}")
    print(f"  改名校:             {cov['rename']}")
    if not report["stage2_applied"]:
        print()
        print("注: 未应用 Stage2 agent 结果（语义匹配待 harness 派发，")
        print("    见 semantic-match/RUN.md）。匹配置为严格匹配口径。")
    if not report["rename_applied"]:
        print("注: 未应用改名 agent 结果（改名配对待 harness 派发，")
        print("    见 research/RUN_RENAME.md）。被删/新增校为上界。")


if __name__ == "__main__":
    main()
