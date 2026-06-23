"""Smoke runner for Slice 6 — rename candidates + edge counts (real data).

Not part of the RED contract (Plan v2: smoke = "被删数 / 改名候选对数 /
新增校/停招校数报告"). Run manually:

    .venv/bin/python -m scripts.run_rename_smoke

It materialises:
  - ``semantic-match/rename_candidates.jsonl`` + ``rename_prompt.md``
    (Task 6.2 — agent dispatch input).
  - ``output/被删旧专业.xlsx`` (Task 6.1 — history majors absent from 2026 at
    schools still present and not renamed).
  - ``output/新增校表.xlsx`` / ``output/停招消失校表.xlsx`` (Task 6.2 — unpaired
    独有校 in each direction).

It does NOT run the agent rename pairing or WebSearch (harness tools).
``学校改名表.xlsx`` / ``特殊情况.xlsx`` are produced after the agent + after
Stage 2 classification respectively; this smoke run reports the candidate
counts that feed them.
"""

from __future__ import annotations

import collections
from pathlib import Path

from openpyxl import load_workbook

from scripts.constants import FLIGHT_BATCH, J3_SHEET
from scripts.rename_detect import prep_rename_candidates, write_rename_prompt
from scripts.stage0_merge import (
    build_dagluben_early,
    build_dagluben_regular,
    build_unified_history,
)
from scripts.stage1_strict import match_strict
from scripts.stage3_edges import deleted_majors, flight_and_special
from scripts.write_edge_tables import (
    write_deleted_major_table,
    write_gone_school_table,
    write_new_school_table,
    write_special_table,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
OUTPUT_DIR = REPO_ROOT / "output"
SEMANTIC_DIR = REPO_ROOT / "semantic-match"


def _load_rows(path: Path, sheet_name: str | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    try:
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def main() -> None:
    j3_path = DATA_DIR / "近三年学校批次专业线差统计.xlsx"
    tq_path = DATA_DIR / "山东省高考提前批录取数据.xlsx"
    dl_path = DATA_DIR / "山东省2026年大绿本招生计划.xlsx"

    j3_rows = _load_rows(j3_path, J3_SHEET)
    tq_rows = _load_rows(tq_path)
    dl_rows = _load_rows(dl_path)

    history = build_unified_history(j3_rows, tq_rows)
    dagluben = [*build_dagluben_regular(dl_rows), *build_dagluben_early(dl_rows)]

    dgl_schools = sorted({d["school"] for d in dagluben if d.get("school")})
    hist_schools = sorted({h["school"] for h in history if h.get("school")})
    dgl_present = set(dgl_schools)

    # --- Task 6.2 rename candidate prep (no agent yet) ----------------------
    candidates = prep_rename_candidates(dgl_schools, hist_schools, topk=5)
    SEMANTIC_DIR.mkdir(parents=True, exist_ok=True)
    write_rename_prompt(candidates, SEMANTIC_DIR)

    # Without the agent run we cannot know which大绿本独有校 are confirmed
    # renamed; for the smoke we treat the rename set as empty so被删 counts
    # reflect the upper bound (真实被删 ⊆ 此数 after agent excludes改名校).
    renamed_dgl: set[str] = set()

    # --- Task 6.1 deleted majors (excludes renamed; here empty) -------------
    # 被删 = 近三年有 + 该校在2026 + 2026 缺该专业。本函数返回「该校在2026 +
    # 非改名校」的近三年历史行；再减去2026 大绿本(校,专业) 对即真被删。
    dgl_school_major = {
        (d.get("school", ""), d.get("major", "")) for d in dagluben
    }
    deleted_pool = deleted_majors(history, dgl_present, renamed_dgl)
    true_deleted = [
        dm for dm in deleted_pool
        if (dm.get("school", ""), dm.get("major", "")) not in dgl_school_major
    ]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_deleted_major_table(true_deleted, OUTPUT_DIR / "被删旧专业.xlsx")

    # --- Task 6.2 unpaired 独有校 (no agent → all 大绿本独有校 are "new") ----
    # 真实跑后 agent 把其中改名者挪入改名表; 剩余即新增校.
    major_count: dict[str, int] = collections.Counter(
        d.get("school", "") for d in dagluben if d.get("school")
    )
    dgl_unique = [s for s in dgl_schools if s not in set(hist_schools)]
    hist_unique = [s for s in hist_schools if s not in dgl_present]

    new_school_rows = [
        {"new_school": s, "major_count_2026": major_count[s]} for s in dgl_unique
    ]
    gone_school_rows = [{"old_school": s} for s in hist_unique]
    write_new_school_table(new_school_rows, OUTPUT_DIR / "新增校表.xlsx")
    write_gone_school_table(gone_school_rows, OUTPUT_DIR / "停招消失校表.xlsx")

    # --- Task 6.1 special: flight-unmatched + remaining unmatched ------------
    # Smoke approximation: strict-match the full dagluben, treat unmatched as
    # the edge pool. Flight rows are the FLIGHT_BATCH majors.
    results = match_strict(dagluben, history)
    unmatched = [
        d for d, r in zip(dagluben, results, strict=True) if not r.get("matched")
    ]
    flight = [d for d in unmatched if d.get("batch") == FLIGHT_BATCH]
    other = [d for d in unmatched if d.get("batch") != FLIGHT_BATCH]
    special = flight_and_special(flight, other)
    write_special_table(special, OUTPUT_DIR / "特殊情况.xlsx")

    # --- Report --------------------------------------------------------------
    print("=== Slice 6 smoke (real data) ===")
    print(f"大绿本独有校数 (rename candidates): {len(dgl_unique)}")
    print(f"  改名候选对数 (topk=5):            {len(candidates)}")
    print(f"历史独有校数 (停招消失候选):        {len(hist_unique)}")
    print(f"被删旧专业数 (改名排除前, 上界):    {len(true_deleted)}")
    print(f"飞行技术(军队)未匹配:               {len(flight)}")
    print(f"剩余未匹配(进特殊表):               {len(other)}")
    print(f"产物: {SEMANTIC_DIR}/rename_candidates.jsonl, rename_prompt.md")
    print(f"      {OUTPUT_DIR}/被删旧专业.xlsx, 新增校表.xlsx, 停招消失校表.xlsx, 特殊情况.xlsx")
    print()
    print("注: 本 smoke 不跑 agent / WebSearch. 学校改名表.xlsx + 改名校专业")
    print("    J/T 留空标记 待 harness 跑完 Task 6.2 agent 后产出.")


if __name__ == "__main__":
    main()
