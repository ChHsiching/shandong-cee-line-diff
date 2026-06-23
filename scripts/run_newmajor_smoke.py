"""Smoke runner for Slice 5 — real-data new-major count + degradation mix.

Not part of the RED contract (Plan v2: smoke = "新增专业数 + 三级分布报告").
Run manually:

    .venv/bin/python -m scripts.run_newmajor_smoke

It materialises ``output/新增专业.xlsx`` and prints the count of identified
新增专业 plus the level-0/1/2 distribution. Estimates use the *strict-unmatched*
pool; a full pipeline run (Slice 7) would feed the post-Stage-2 unmatched set.
Here we approximate by treating every大绿本 row that strict-match missed and
that has no same-school core candidate as a new major — enough signal to
verify the degradation chain on real data.
"""

from __future__ import annotations

import collections
from pathlib import Path

from openpyxl import load_workbook

from scripts.constants import (
    J3_SHEET,
)
from scripts.stage0_merge import build_dagluben_early, build_dagluben_regular, build_unified_history
from scripts.stage1_strict import match_strict
from scripts.stage3_newmajor import estimate
from scripts.write_edge_tables import (
    identify_new_majors,
    write_new_major_table,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
OUTPUT_DIR = REPO_ROOT / "output"


def _load_rows(path: Path, sheet_name: str | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    try:
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def main() -> None:
    j3_path = DATA_DIR / "近三年学校批次专业线差统计.xlsx"
    tq_path = DATA_DIR / "山东省高考提前批录取数据.xlsx"
    dl_path = DATA_DIR / "山东省2026年大绿本招生计划.xlsx"

    j3_rows = _load_rows(j3_path, J3_SHEET)
    tq_rows = _load_rows(tq_path)
    dl_rows = _load_rows(dl_path)

    history = build_unified_history(j3_rows, tq_rows)
    dagluben = [*build_dagluben_regular(dl_rows), *build_dagluben_early(dl_rows)]

    results = match_strict(dagluben, history)
    unmatched_dagluben = [
        d for d, r in zip(dagluben, results, strict=True) if not r.get("matched")
    ]

    new_majors = identify_new_majors(unmatched_dagluben, history)

    # Group same-school history once for the estimator.
    by_school: dict[str, list] = collections.defaultdict(list)
    for h in history:
        by_school[h["school"]].append(h)

    table_rows = []
    level_counts = collections.Counter()
    for d in new_majors:
        est = estimate(d, by_school.get(d["school"], []))
        level_counts[est["level"]] += 1
        table_rows.append({
            "school": d.get("school", ""),
            "major": d.get("major", ""),
            "subject": d.get("subject", ""),
            "value": est.get("value"),
            "level": est.get("level"),
            "n": est.get("n"),
            "log": est.get("log", ""),
        })

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "新增专业.xlsx"
    write_new_major_table(table_rows, out_path)

    total = len(new_majors)
    print(f"新增专业总数: {total}")
    print(f"  退化0 (同校同选科均值):  {level_counts[0]}")
    print(f"  退化1 (同校全专业均值):  {level_counts[1]}")
    print(f"  退化2 (整校无历史):      {level_counts[2]}")
    print(f"产出: {out_path}")


if __name__ == "__main__":
    main()
