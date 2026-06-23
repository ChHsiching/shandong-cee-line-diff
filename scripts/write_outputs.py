"""Output writers for the admission-data pipeline.

Two products (spec §7):
    - Hierarchical (大绿本_附线差_分层版.xlsx): a full copy of the大绿本
      workbook with seven columns appended at the row end
      (近三年统计线差 / 近三年线差标准差 / 匹配阶段 / 单年数据 / 选科漂移 /
      复核结果 / 原因备注). Every original row is preserved verbatim; non-major
      rows leave the seven new cells blank. Original columns are never
      overwritten.
    - Flat (大绿本_附线差_扁平版.xlsx): only专业行, each with all original
      fields plus the seven appended columns.

iteration-3 (structured-columns): the legacy single「匹配日志」cell is gone;
the seven row-end columns are ``[J, T] + list(split_log(log).values())`` so
users can filter directly by stage / single-year / drift / verify (spec §2.1).

``write_edge_tables.py`` handles the boundary tables (Slice 5/6) so this
module's file domain is stable across slices (Plan v2 binding).
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl

from scripts.constants import LOG_ZHUANKE_OUT_OF_SCOPE
from scripts.models import MatchResult
from scripts.structured_log import split_log

__all__ = [
    "write_hierarchical",
    "write_flat",
    "COL_J",
    "COL_T",
    "COL_STAGE",
    "COL_SINGLE_YEAR",
    "COL_DRIFT",
    "COL_VERIFIED",
    "COL_NOTE",
    "HEADER_J",
    "HEADER_T",
    "HEADER_STAGE",
    "HEADER_SINGLE_YEAR",
    "HEADER_DRIFT",
    "HEADER_VERIFIED",
    "HEADER_NOTE",
]

# Append-position header labels (Plan v2 修订 binding: 7 named constants,
# fixed order). Row-end layout = [J, T, 匹配阶段, 单年数据, 选科漂移,
# 复核结果, 原因备注].
HEADER_J = "近三年统计线差"
HEADER_T = "近三年线差标准差"
HEADER_STAGE = "匹配阶段"
HEADER_SINGLE_YEAR = "单年数据"
HEADER_DRIFT = "选科漂移"
HEADER_VERIFIED = "复核结果"
HEADER_NOTE = "原因备注"

# 1-based column indices for the appended seven, given the大绿本 has 12 cols.
COL_J = 13
COL_T = 14
COL_STAGE = 15
COL_SINGLE_YEAR = 16
COL_DRIFT = 17
COL_VERIFIED = 18
COL_NOTE = 19

# 大绿本 column where 代号(E) and 名称(F) live (1-based); both non-empty on
# 专业行.
COL_CODE = 5
COL_NAME = 6


def _is_major_row(row_cells) -> bool:
    code = row_cells[COL_CODE - 1] if len(row_cells) >= COL_CODE else None
    name = row_cells[COL_NAME - 1] if len(row_cells) >= COL_NAME else None
    return code not in (None, "") and name not in (None, "")


def _index_by_src_row(results: Iterable[MatchResult]) -> dict[int, MatchResult]:
    out: dict[int, MatchResult] = {}
    for r in results:
        idx = r.get("src_row_idx", 0)
        if idx and idx not in out:
            out[idx] = r
    return out


def _open_template(path: str | Path) -> openpyxl.Workbook:
    """Open the source workbook for copying. write_outputs must not mutate the
    source, so we load it without read_only (we need to append cells) and save
    to a *different* path."""
    return openpyxl.load_workbook(Path(path), data_only=True)


def _write_row_end(ws, row_idx: int, res: MatchResult) -> None:
    """Write the 7 row-end cells for a MatchResult.

    Layout: [J, T, 匹配阶段, 单年数据, 选科漂移, 复核结果, 原因备注]
    where the last 5 = ``list(split_log(res.log).values())``.
    """
    structured = split_log(res.get("log") or "")
    ws.cell(row=row_idx, column=COL_J, value=res.get("J"))
    ws.cell(row=row_idx, column=COL_T, value=res.get("T"))
    ws.cell(row=row_idx, column=COL_STAGE, value=structured["匹配阶段"])
    ws.cell(row=row_idx, column=COL_SINGLE_YEAR, value=structured["单年数据"])
    ws.cell(row=row_idx, column=COL_DRIFT, value=structured["选科漂移"])
    ws.cell(row=row_idx, column=COL_VERIFIED, value=structured["复核结果"])
    ws.cell(row=row_idx, column=COL_NOTE, value=structured["原因备注"])


def _write_header_row_end(ws, row_idx: int) -> None:
    """Write the 7 row-end header labels."""
    ws.cell(row=row_idx, column=COL_J, value=HEADER_J)
    ws.cell(row=row_idx, column=COL_T, value=HEADER_T)
    ws.cell(row=row_idx, column=COL_STAGE, value=HEADER_STAGE)
    ws.cell(row=row_idx, column=COL_SINGLE_YEAR, value=HEADER_SINGLE_YEAR)
    ws.cell(row=row_idx, column=COL_DRIFT, value=HEADER_DRIFT)
    ws.cell(row=row_idx, column=COL_VERIFIED, value=HEADER_VERIFIED)
    ws.cell(row=row_idx, column=COL_NOTE, value=HEADER_NOTE)


def write_hierarchical(
    src_path: str | Path,
    results: Iterable[MatchResult],
    out_path: str | Path,
) -> None:
    """Copy the大绿本 workbook verbatim and append the 7 row-end cells on
    matched major rows.

    - Every original row (header / 批次头 / 小标题 / 学校行 / 专业行) preserved.
    - Original columns (1-12) never overwritten.
    - The seven new columns (13-19) are added; only专业行 that have a
      :class:`MatchResult` carry values, all others stay blank. 专科 专业行
      without a result are annotated with LOG_ZHUANKE_OUT_OF_SCOPE so
      ``匹配阶段`` becomes ``专科（超范围）`` (spec §3).
    """
    wb = _open_template(src_path)
    try:
        ws = wb.active
        _write_header_row_end(ws, row_idx=1)

        results_by_idx = _index_by_src_row(results)

        # Iterate over rows; row index in openpyxl is 1-based and matches
        # src_row_idx (header is row 1).
        for row_idx in range(2, ws.max_row + 1):
            # Defensive: only act on 专业行 (代号+名称 非空).
            code = ws.cell(row=row_idx, column=COL_CODE).value
            name = ws.cell(row=row_idx, column=COL_NAME).value
            if code in (None, "") or name in (None, ""):
                continue
            subtitle = ws.cell(row=row_idx, column=2).value
            res = results_by_idx.get(row_idx)
            if res is None:
                # 专业行 无结果 = 专科（本科行经 Fix B 均有 MatchResult）。
                # 标注超范围 → split_log turns it into 匹配阶段=专科（超范围）.
                if "专科" in str(subtitle or ""):
                    structured = split_log(LOG_ZHUANKE_OUT_OF_SCOPE)
                    ws.cell(row=row_idx, column=COL_STAGE,
                            value=structured["匹配阶段"])
                    # J/T + other structured cells stay blank (专科 excluded).
                continue
            _write_row_end(ws, row_idx, res)

        out_p = Path(out_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_p)
    finally:
        wb.close()


def write_flat(
    src_path: str | Path,
    results: Iterable[MatchResult],
    out_path: str | Path,
) -> None:
    """Write a flat table containing only专业行 + the seven appended columns.

    Columns 1-12 carry the original大绿本 fields; 13-19 carry
    ``[J, T, 匹配阶段, 单年数据, 选科漂移, 复核结果, 原因备注]``. Non-major rows
    (批次头/小标题/学校行) are omitted entirely. 专科 专业行 are still excluded
    (spec §3, unchanged).
    """
    src_wb = _open_template(src_path)
    out_wb = openpyxl.Workbook()
    try:
        src_ws = src_wb.active
        out_ws = out_wb.active
        out_ws.title = src_ws.title

        results_by_idx = _index_by_src_row(results)

        # Header row: original 12 columns + 7 row-end labels.
        out_row = 1
        for col_idx in range(1, 13):
            out_ws.cell(row=out_row, column=col_idx,
                        value=src_ws.cell(row=1, column=col_idx).value)
        _write_header_row_end(out_ws, row_idx=out_row)
        out_row += 1

        for src_row_idx in range(2, src_ws.max_row + 1):
            # Read original row cells.
            cells = [
                src_ws.cell(row=src_row_idx, column=c).value
                for c in range(1, 13)
            ]
            if not _is_major_row(cells):
                continue
            # 专科 专业行不在本次整理范围（仅本科），扁平版剔除。
            if "专科" in str(cells[1] or ""):
                continue
            for col_idx, val in enumerate(cells, start=1):
                out_ws.cell(row=out_row, column=col_idx, value=val)
            res = results_by_idx.get(src_row_idx)
            if res is not None:
                _write_row_end(out_ws, out_row, res)
            else:
                # Defensive: blank the 7 row-end cells.
                for col in (COL_J, COL_T, COL_STAGE, COL_SINGLE_YEAR,
                            COL_DRIFT, COL_VERIFIED, COL_NOTE):
                    out_ws.cell(row=out_row, column=col, value=None)
            out_row += 1

        out_p = Path(out_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)
        out_wb.save(out_p)
    finally:
        src_wb.close()
        out_wb.close()
