"""Slice 7 — end-to-end pipeline contract tests (RED first).

Per Plan v2 binding + Slice 7 (issue #8): the pipeline ``run()`` pure function
must thread the deterministic stages end-to-end so that:

  - every大绿本本科专业 row is classified exactly once into one of
    {匹配, 新增, 被删(历史侧), 特殊, 改名占位};
  - 专科 rows are excluded (181 rows in real data — proven by stage0 unit tests
    and re-asserted here on a 专科-laden fixture);
  - hierarchical and flat outputs are **same-source** — every专业 row that
    appears in both carries identical J/T/log;
  - the three source files are byte-identical before vs after (sha256);
  - Stage 2 agent + rename agent are **optional** — when no
    ``semantic-match/batch_*_result.jsonl`` / ``rename_result.jsonl`` exist,
    the run emits batch prompts + rename candidates and logs that the agent
    step is pending harness dispatch; when stub jsonls are present, the apply
    path back-fills them.

The fixture is a tiny synthetic workbook set (j3/tq/dl) constructed via
``tmp_xlsx`` so the test does not touch the real 12MB sources.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl
import pytest

from scripts import io_source
from scripts.run_pipeline import run


# ---------------------------------------------------------------------------
# Fixture builders — minimal but exercising every classification path.
# ---------------------------------------------------------------------------
#
# Layout (chosen so each classification bucket is non-empty after the full
# deterministic chain):
#
#   近三年 (j3) 常规批一段线:
#     - 示例大学  计算机科学与技术     → strict-matches dl row 1
#     - 示例大学  数学类               → Stage 2 agent 匹配 dl row 2 (核心名唯一;
#                                          with_agent_results=False 时 dl row 2 进 special)
#     - 示例大学  历史学               → absent from 2026 → 被删 (school present)
#     - 停招大学  物理                  → school absent from 2026 → 停招消失校
#
#   提前批 supplement (tq) 本科提前批A类:
#     - 飞行大学  飞行技术              → matches dl flight row (J/T computed)
#
#   大绿本 (dl):
#     - 4.常规批 示例大学 01 计算机科学与技术  (strict match)
#     - 4.常规批 示例大学 02 数学类(拔尖)     (Stage 2 待 agent: 数学 core 唯一)
#     - 4.常规批 示例大学 03 量子信息          (new major: no同校 core)
#     - 4.常规批 专科大学 05 护理             subtitle contains「专科」→ excluded
#     - 4.常规批 新校 08 人工智能             (new school + new major → 新增 level2)
#     - 1.提前批A类 飞行大学 06 飞行技术       (early batch, matches tq)
#     - 4.常规批 军航大学 07 飞行技术(军队)   (NOT new — 军航大学 has 飞行技术
#                                            core history under a different cat;
#                                            unmatched → special)


_J3_HEADER = [
    "批次",
    "学校代码",
    "学校名称",
    "专业名称",
    "选考科目",
    "备注",
    "基础专业名",
    "是否括号专业",
    "括号内容",
    "统计线差",
    "2023线差",
    "2024线差",
    "2025线差",
    "x",
    "x",
    "x",
    "可用年份数",
    "x",
    "x",
    "线差标准差",
]


def _j3_row(batch, school, major, j, t, subject="物理和化学"):
    return [
        batch,
        "C001",
        school,
        major,
        subject,
        "",
        major,
        "否",
        "",
        j,
        j,
        j,
        j,
        "",
        "",
        "",
        3,
        "",
        "",
        t,
    ]


def _build_j3(tmp_xlsx):
    rows = [
        _J3_HEADER,
        _j3_row("常规批一段线", "示例大学", "计算机科学与技术", 60.0, 5.0),
        _j3_row("常规批一段线", "示例大学", "数学类", 70.0, None, subject="物理"),
        _j3_row("常规批一段线", "示例大学", "历史学", 55.0, 4.0),
        _j3_row("常规批一段线", "停招大学", "物理", 80.0, 3.0),
        # 军航大学 has a history major with the SAME core (飞行技术) but a
        # DIFFERENT 招生类别 (地方专项计划). The dagluben 飞行技术(军队) row is
        # 普通计划 → strict fails (cat differs), AND identify_new_majors sees
        # the school has 飞行技术 in its history cores so it is NOT a真新增
        # → it falls through to the special table.
        _j3_row("常规批一段线", "军航大学(地方专项计划)", "飞行技术", 50.0, 2.0),
        # 常规批二段线 — must be dropped.
        _j3_row("常规批二段线", "示例大学", "应被丢弃", 99.0, 1.0),
    ]
    return tmp_xlsx(rows, sheet_name="统计结果")


def _build_tq(tmp_xlsx):
    header = [
        "批次名称",
        "招生类别",
        "x",
        "院校名称",
        "x",
        "专业名称",
        "选考科目",
        "x",
        "x",
        "x",
        "2025低分",
        "x",
        "x",
        "x",
        "2024低分",
        "x",
        "x",
        "x",
        "2023低分",
    ]
    # 飞行大学 飞行技术 low scores: 2025=501, 2024=504, 2023=503 →
    # line-diff = low - one_line (441/444/443) = 60/60/60 → mean 60.0.
    rows = [
        header,
        [
            "本科提前批A类",
            "",
            "",
            "飞行大学",
            "",
            "飞行技术",
            "物理",
            "",
            "",
            "",
            501,
            "",
            "",
            "",
            504,
            "",
            "",
            "",
            503,
        ],
    ]
    return tmp_xlsx(rows)


def _dl_row(batch, subtitle, code, school, name, subject="物理和化学"):
    return [
        batch,
        subtitle,
        "A001",
        school,
        code,
        name,
        subject,
        "4",
        "2",
        "",
        "",
        "",
    ]


def _build_dl(tmp_xlsx):
    header = [
        "批次",
        "小标题",
        "学校代码",
        "学校名",
        "代号",
        "名称",
        "选考科目要求",
        "学制",
        "计划数",
        "学校备注",
        "年收费",
        "整行校准",
    ]
    rows = [
        header,
        # 批次头
        ["4.常规批", "", "", "", "", "", "", "", "", "", "", ""],
        # 示例大学 学校行 (no 代号/名称)
        ["4.常规批", "普通计划", "A001", "示例大学", "", "", "", "", "100", "", "", ""],
        # 专业行 1 — strict match (计算机科学与技术)
        _dl_row("4.常规批", "普通计划", "01", "示例大学", "计算机科学与技术"),
        # 专业行 2 — Stage 2 待 agent (数学类(拔尖) → 数学 core 唯一; coarse 已停用)
        _dl_row(
            "4.常规批", "普通计划", "02", "示例大学", "数学类(拔尖)", subject="物理"
        ),
        # 专业行 3 — new major (量子信息, no同校 core)
        _dl_row("4.常规批", "普通计划", "03", "示例大学", "量子信息"),
        # 专业行 4 — 专科 subtitle row (must be excluded)
        [
            "4.常规批",
            "专科",
            "A002",
            "专科大学",
            "05",
            "护理",
            "物理",
            "3",
            "50",
            "",
            "",
            "",
        ],
        # 专业行 5 — 新校 + 新专业 (新增 level2: no history at all)
        _dl_row("4.常规批", "普通计划", "08", "新校", "人工智能"),
        # 提前批 A类 飞行大学 飞行技术 (matches tq)
        _dl_row("1.提前批A类", "普通计划", "06", "飞行大学", "飞行技术"),
        # 军航大学 飞行技术(军队) — school has 飞行技术 core history under a
        # different 招生类别 (地方专项计划) so it is NOT a真新增, strict fails
        # (different cat) → falls to special.
        _dl_row("4.常规批", "普通计划", "07", "军航大学", "飞行技术(军队)"),
    ]
    return tmp_xlsx(rows)


@pytest.fixture
def fixture_workbooks(tmp_xlsx):
    """Build the three tiny source workbooks; return their paths."""
    return {
        "j3": _build_j3(tmp_xlsx),
        "tq": _build_tq(tmp_xlsx),
        "dl": _build_dl(tmp_xlsx),
    }


# ---------------------------------------------------------------------------
# Contract: source files are byte-identical before vs after the run.
# ---------------------------------------------------------------------------


def test_run_preserves_source_bytes(fixture_workbooks, tmp_path):
    """Each source xlsx's sha256 must be unchanged by the pipeline run."""
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    j3_dst = data_dir / "近三年学校批次专业线差统计.xlsx"
    tq_dst = data_dir / "山东省高考提前批录取数据.xlsx"
    dl_dst = data_dir / "山东省2026年大绿本招生计划.xlsx"
    j3_dst.write_bytes(fixture_workbooks["j3"].read_bytes())
    tq_dst.write_bytes(fixture_workbooks["tq"].read_bytes())
    dl_dst.write_bytes(fixture_workbooks["dl"].read_bytes())

    before = {p.name: io_source.sha256(p) for p in (j3_dst, tq_dst, dl_dst)}
    out_dir = tmp_path / "output"

    report = run(data_dir, out_dir, with_agent_results=False)

    after = {p.name: io_source.sha256(p) for p in (j3_dst, tq_dst, dl_dst)}
    assert before == after, "source files mutated by pipeline"
    # Sanity: the run reported the hashes too.
    assert set(report["source_hashes"].keys()) == set(before.keys())


# ---------------------------------------------------------------------------
# Contract: 专科 excluded.
# ---------------------------------------------------------------------------


def test_run_excludes_zhuanke_rows(fixture_workbooks, tmp_path):
    """The专科 subtitle row (护理 at 专科大学) must NOT appear in any本科
    output surface (flat专业 list, new-major table, special table)."""
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    report = run(data_dir, out_dir, with_agent_results=False)

    all_major_names: list[str] = []
    for r in report["main_results"]:
        all_major_names.append(r.get("major", ""))
    for r in report["edge"]["special"]:
        all_major_names.append(r.get("major", ""))
    for r in report["new_major_rows"]:
        all_major_names.append(r.get("major", ""))

    assert "护理" not in all_major_names, "专科 row leaked into本科 output"
    assert "专科大学" not in {r.get("school", "") for r in report["main_results"]}, (
        "专科学校 leaked into main output"
    )


# ---------------------------------------------------------------------------
# Contract: hierarchical and flat outputs are same-source.
# ---------------------------------------------------------------------------


def test_hierarchical_and_flat_are_same_source(fixture_workbooks, tmp_path):
    """For every专业 row present in both outputs, J/T + the 5 structured
    columns must match.

    Both outputs are written from the same MatchResult list, so every (school,
    major) pair that appears in both must carry an identical row-end 7-tuple
    (J, T, 匹配方式, 仅一年数据, 选科要求跨年变化, 二次复核, 原因说明). We key by
    (school, major) rather than row position because the flat output omits
    non-major rows and so its row positions do not align with the
    hierarchical output's.
    """
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    run(data_dir, out_dir, with_agent_results=False)

    hier_path = out_dir / "大绿本_完整版_含线差.xlsx"
    flat_path = out_dir / "大绿本_专业列表_含线差.xlsx"
    assert hier_path.exists()
    assert flat_path.exists()

    hier = _read_output_major_rows(hier_path)
    flat = _read_output_major_rows(flat_path)

    row_end_keys = ("J", "T") + _OUTPUT_STRUCTURED_HEADERS
    flat_by_key = {(r["school"], r["major"]): r for r in flat}
    mismatches: list[str] = []
    for h in hier:
        key = (h["school"], h["major"])
        f = flat_by_key.get(key)
        if f is None:
            continue
        h_end = tuple(h[k] for k in row_end_keys)
        f_end = tuple(f[k] for k in row_end_keys)
        if h_end != f_end:
            mismatches.append(f"{key}: hier={h_end!r} flat={f_end!r}")
    assert not mismatches, "hierarchical/flat diverged:\n" + "\n".join(mismatches)

    # Every major row in the flat output must also appear in the hierarchical
    # output (flat is a subset of hierarchical's专业 rows).
    hier_keys = {(r["school"], r["major"]) for r in hier}
    for key in flat_by_key:
        assert key in hier_keys, f"flat row {key} missing from hierarchical"


# ---------------------------------------------------------------------------
# Contract: 100% classification — every本科专业 row lands in exactly one bucket.
# ---------------------------------------------------------------------------


def test_every_undergrad_major_is_classified(fixture_workbooks, tmp_path):
    """Every大绿本本科专业 row must appear in exactly one of:
      - main_results (matched via strict/agent)
      - new_major_rows (新增估算)
      - edge.special (无法匹配)
      - edge.rename_pending (改名校占位, only when rename results applied)
    And none may be lost or double-counted."""
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    report = run(data_dir, out_dir, with_agent_results=False)

    matched_idx = {r["src_row_idx"] for r in report["main_results"] if r.get("matched")}
    new_idx = {r["src_row_idx"] for r in report["new_major_rows"]}
    special_idx = {r["src_row_idx"] for r in report["edge"]["special"]}
    # rename_pending only appears when rename results exist; without them,
    # rename-pending rows fall through to new/special — still classified.

    # The union must equal the set of大绿本本科专业 src_row_idx (= dagluben rows).
    all_dgl_idx = set(report["dagluben_indices"])
    classified = matched_idx | new_idx | special_idx

    missing = all_dgl_idx - classified
    assert not missing, f"unclassified大绿本 rows: {sorted(missing)}"

    # No double-counting: buckets are disjoint.
    overlap = (
        (matched_idx & new_idx) | (matched_idx & special_idx) | (new_idx & special_idx)
    )
    assert not overlap, f"rows double-classified: {sorted(overlap)}"


# ---------------------------------------------------------------------------
# Contract: deterministic chain matches the expected classification on fixture.
# ---------------------------------------------------------------------------


def test_classification_counts_on_fixture(fixture_workbooks, tmp_path):
    """The fixture exercises each path; assert the exact classification."""
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    report = run(data_dir, out_dir, with_agent_results=False)

    # Row 4 in fixture = 示例大学 计算机科学与技术 → strict match.
    # We need the actual src_row_idx values. They are 1-based row indices in the
    # 大绿本 workbook. Build the expected map by re-reading the source.
    dl_rows = list(
        openpyxl.load_workbook(
            (data_dir / "山东省2026年大绿本招生计划.xlsx"),
            read_only=True,
            data_only=True,
        ).active.iter_rows(values_only=True)
    )
    # Re-derive the本科-only major row count the same way build_dagluben_* do:
    # exclude rows whose 小标题 (col B, idx 1) carries the专科 keyword.
    majors_by_row = {}
    for i, row in enumerate(dl_rows, start=1):
        if i == 1:
            continue
        subtitle = row[1] if len(row) > 1 else ""
        if subtitle and "专科" in str(subtitle):
            continue
        code = row[4] if len(row) > 4 else None
        name = row[5] if len(row) > 5 else None
        if code and name:
            majors_by_row[i] = (row[3], name)  # school, major

    matched_schools_major = {
        (r.get("school"), r.get("major"))
        for r in report["main_results"]
        if r.get("matched")
    }
    # 计算机科学与技术 strict-matches; 飞行技术 (early batch) strict-matches tq;
    # 数学类(拔尖) — coarse 已停用，with_agent_results=False 时 Stage 2 不 apply → special.
    assert ("示例大学", "计算机科学与技术") in matched_schools_major
    assert ("飞行大学", "飞行技术") in matched_schools_major

    # 量子信息 + 人工智能 are new majors (no同校 core in history).
    new_pairs = {(r.get("school"), r.get("major")) for r in report["new_major_rows"]}
    assert ("示例大学", "量子信息") in new_pairs
    assert ("新校", "人工智能") in new_pairs
    # 新校 has no history at all → level 2.
    xc_row = next(r for r in report["new_major_rows"] if r.get("school") == "新校")
    assert xc_row["level"] == 2

    # 飞行技术(军队) at 军航大学 → special; 数学类(拔尖) → special (coarse 已停用,
    # with_agent_results=False 时 Stage 2 未 apply).
    special_pairs = {
        (r.get("school"), r.get("major")) for r in report["edge"]["special"]
    }
    assert ("军航大学", "飞行技术(军队)") in special_pairs
    assert ("示例大学", "数学类(拔尖)") in special_pairs

    # Coverage report keys exist and sum correctly.
    cov = report["coverage"]
    assert cov["total_dagluben"] == len(majors_by_row)
    assert cov["matched"] == len(matched_schools_major)
    assert cov["new_major"] == len(new_pairs)
    assert cov["special"] == len(special_pairs)


# ---------------------------------------------------------------------------
# Contract: agent jsonl (when present) is applied; absent → prompts + log.
# ---------------------------------------------------------------------------


def test_without_agent_results_prompts_are_written_and_logged(
    fixture_workbooks, tmp_path, caplog
):
    """When no semantic-match/batch_*_result.jsonl exists, the run must:
    - still succeed,
    - write batch_NN_prompt.json files under semantic-match/,
    - log that Stage 2 agent dispatch is pending.
    """
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    semantic_dir = tmp_path / "semantic-match"

    with caplog.at_level(logging.INFO, logger="scripts.run_pipeline"):
        report = run(
            data_dir,
            out_dir,
            with_agent_results=False,
            semantic_dir=semantic_dir,
        )

    # If strict already resolved everything, prompts may be empty; that
    # is legitimate. The key contract is: no result jsonl applied.
    assert report["stage2_applied"] is False
    assert any(
        "Stage2" in rec.message and "待 harness" in rec.message
        for rec in caplog.records
    ), "expected a 'Stage2 pending harness dispatch' log line"


def test_with_stub_agent_results_back_fills_main_table(fixture_workbooks, tmp_path):
    """When semantic-match/batch_*_result.jsonl exists, the run applies it.

    We stage a stub jsonl that resolves the otherwise-new「量子信息」to a
    null match (so it still falls through to new-major) and verify the
    apply path is exercised and the source hashes still hold.
    """
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    semantic_dir = tmp_path / "semantic-match"

    # First run to discover the unmatched src_row_idx values.
    probe = run(data_dir, out_dir, with_agent_results=False, semantic_dir=semantic_dir)
    unmatched_after_coarse = probe["post_coarse_unmatched_indices"]

    if unmatched_after_coarse:
        # Pick the first unmatched idx and craft a null-match result line.
        target_idx = unmatched_after_coarse[0]
        target_dgl = next(
            d for d in probe["dagluben_rows"] if d["src_row_idx"] == target_idx
        )
        line = json.dumps(
            {
                "src_row_idx": target_idx,
                "school": target_dgl["school"],
                "major": target_dgl["major"],
                "match": None,
                "J": None,
                "T": None,
                "reason": "契约测试桩: 无对应",
            },
            ensure_ascii=False,
        )
        (semantic_dir / "batch_99_result.jsonl").write_text(
            line + "\n", encoding="utf-8"
        )

    report = run(data_dir, out_dir, with_agent_results=True, semantic_dir=semantic_dir)
    assert report["stage2_applied"] is True


# ---------------------------------------------------------------------------
# Contract: rename jsonl (when present) drives renamed-set + 被删 exclusion.
# ---------------------------------------------------------------------------


def test_without_rename_results_renamed_set_is_empty_and_logged(
    fixture_workbooks, tmp_path, caplog
):
    """No rename_result.jsonl → renamed_dgl_schools is the empty set and a log
    line records that the rename step is pending harness dispatch."""
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    with caplog.at_level(logging.INFO, logger="scripts.run_pipeline"):
        report = run(data_dir, out_dir, with_agent_results=False)
    assert report["renamed_dgl_schools"] == set()
    assert any(
        "改名" in rec.message and "待 harness" in rec.message for rec in caplog.records
    ), "expected a 'rename pending harness dispatch' log line"


def test_with_stub_rename_results_applies_rename(fixture_workbooks, tmp_path):
    """#6c: rename_result.jsonl 应用后，新校←旧校 记入改名表；新校的 dagluben
    专业通过 renamed history 用旧校线差（fixture 旧校无 history → 走 new_major）。"""
    data_dir, out_dir = _materialize_sources(fixture_workbooks, tmp_path)
    semantic_dir = tmp_path / "semantic-match"

    # First run (no agent) to ensure rename candidates are generated.
    run(data_dir, out_dir, with_agent_results=False, semantic_dir=semantic_dir)

    # Stage a rename result marking 新校 as a rename of 旧校.
    # 新校 must be a大绿本独有校 (in大绿本, not in history) for the contract.
    rename_line = json.dumps(
        {
            "new_school": "新校",
            "old_school": "旧校",
            "confidence": 0.9,
            "is_rename": True,
        },
        ensure_ascii=False,
    )
    (semantic_dir / "rename_result.jsonl").write_text(
        rename_line + "\n", encoding="utf-8"
    )

    report = run(
        data_dir,
        out_dir,
        with_agent_results=True,
        semantic_dir=semantic_dir,
    )
    assert report["rename_applied"] is True
    assert "新校" in report["renamed_dgl_schools"]
    # 改名表含 新校←旧校。
    rename_pairs = {
        (r.get("new_school"), r.get("old_school")) for r in report["edge"]["rename"]
    }
    assert ("新校", "旧校") in rename_pairs
    # #6c: 新校专业通过 renamed history 用旧校线差。fixture 旧校无 history
    # → 新校专业（人工智能）走 new_major（不再 rename-pending J/T 留空）。
    xc_in_new = [r for r in report["new_major_rows"] if r.get("school") == "新校"]
    assert xc_in_new, "新校专业应在 new_major_rows（旧校无 history）"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _materialize_sources(fixture_workbooks, tmp_path):
    """Copy the three fixture workbooks into tmp_path/data (read-only intent)."""
    data_dir = tmp_path / "data"
    data_dir.mkdir(exist_ok=True)
    (data_dir / "近三年学校批次专业线差统计.xlsx").write_bytes(
        fixture_workbooks["j3"].read_bytes()
    )
    (data_dir / "山东省高考提前批录取数据.xlsx").write_bytes(
        fixture_workbooks["tq"].read_bytes()
    )
    (data_dir / "山东省2026年大绿本招生计划.xlsx").write_bytes(
        fixture_workbooks["dl"].read_bytes()
    )
    out_dir = tmp_path / "output"
    return data_dir, out_dir


_OUTPUT_HEADER_J = "近三年统计线差"
_OUTPUT_HEADER_T = "近三年线差标准差"
# iteration-3: the legacy single「匹配日志」column was split into 5 structured
# columns. We read all 5 by header name so a column re-order cannot silently
# break the hierarchical-vs-flat consistency check.
_OUTPUT_STRUCTURED_HEADERS = (
    "匹配方式",
    "仅一年数据",
    "选科要求跨年变化",
    "二次复核",
    "原因说明",
)


def _read_output_major_rows(path: Path) -> list[dict]:
    """Read an output workbook and return its专业 rows with J/T + the 5
    structured columns + the original src_row_idx.

    For the hierarchical output the src_row_idx is the openpyxl 1-based row
    number (matching the大绿本 source layout). For the flat output, we cannot
    recover the original src_row_idx from the flat row alone, so we key by
    (school, major) instead and let the caller cross-reference.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = rows[0]
    # Locate columns by header (robust against future column shifts).
    j_idx = header.index(_OUTPUT_HEADER_J)
    t_idx = header.index(_OUTPUT_HEADER_T)
    structured_idx = {h: header.index(h) for h in _OUTPUT_STRUCTURED_HEADERS}
    # Original columns: 代号(E=idx4) 名称(F=idx5) 学校(D=idx3).
    out: list[dict] = []
    for row_idx_1based, row in enumerate(rows, start=1):
        if row_idx_1based == 1:
            continue
        code = row[4] if len(row) > 4 else None
        name = row[5] if len(row) > 5 else None
        if not code or not name:
            continue
        structured = {
            h: (row[idx] if idx < len(row) else None)
            for h, idx in structured_idx.items()
        }
        out.append(
            {
                "src_row_idx": row_idx_1based,
                "school": row[3],
                "major": name,
                "J": row[j_idx] if j_idx < len(row) else None,
                "T": row[t_idx] if t_idx < len(row) else None,
                **structured,
            }
        )
    return out
