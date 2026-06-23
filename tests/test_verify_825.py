"""Tests for scripts.verify_825 — 825弃用前重叠验证 (spec §6 Stage 0).

契约: 近三年 `提前批` (825 rows, being deprecated) 的 (schoolcode, nfk(majorname))
键应全部 ∈ 补充表 (本科 A+B) 的同键集. 独有行写入 intermediate/s2_j3_early_only.csv
供人工核验, 不静默丢 (spec §9 日志: `独有，待人工核验`).

Per Plan v2 this is a **契约测试 (不硬 FAIL)**: assert reported_count == len(csv
独有行); the独有 rows are surfaced for human review rather than blocking CI.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Sequence

import pytest

from scripts import verify_825


def _j3_pairs(rows: Iterable[Sequence]) -> list[tuple[str, str]]:
    """Build (schoolcode, nfk(majorname)) pairs from raw近三年 rows (header incl)."""
    pairs: list[tuple[str, str]] = []
    for r in rows:
        if not r or r[0] in (None, "batch", "批次"):
            continue
        if r[0] == "提前批":
            pairs.append((str(r[1]), verify_825.nfk(r[3])))
    return pairs


def _tq_pairs(rows: Iterable[Sequence]) -> list[tuple[str, str]]:
    """Build (schoolcode, nfk(majorname)) pairs from raw supplement rows (本科 only)."""
    pairs: list[tuple[str, str]] = []
    for r in rows:
        if not r or r[0] not in ("本科提前批A类", "本科提前批B类"):
            continue
        pairs.append((str(r[2]), verify_825.nfk(r[5])))
    return pairs


# --- pure function: pair extraction ----------------------------------------

def test_extract_j3_early_pairs_filters_only_early_batch():
    rows = [
        ("批次", "code", "school", "major", "subject"),
        ("常规批一段线", "A001", "X大学", "数学", "物理"),
        ("提前批", "B002", "Y大学", "小语种", "历史"),
        ("提前批", "B003", "Z大学", "治安学", "历史"),
    ]
    pairs = verify_825.extract_j3_early_pairs(rows)
    assert pairs == [("B002", "小语种"), ("B003", "治安学")]


def test_extract_tq_benke_pairs_filters_only_benke_a_b():
    rows = [
        ("批次名称", "招生类别", "院校代码", "院校名称", "专业代码", "专业名称"),
        ("本科提前批A类", "军事类", "P002", "国防科技大学", "01", "软件工程"),
        ("本科提前批B类", "公安政法类", "P010", "公安大学", "02", "治安学"),
        ("专科提前批", "其他类", "Z001", "某职业学院", "03", "护理"),
    ]
    pairs = verify_825.extract_tq_benke_pairs(rows)
    assert pairs == [("P002", "软件工程"), ("P010", "治安学")]


def test_extract_pairs_applies_nfk_to_majorname():
    rows = [
        ("批次名称", "招生类别", "院校代码", "院校名称", "专业代码", "专业名称"),
        ("本科提前批A类", "军事类", "P002", "国防科技大学", "01", "软件工程（男）"),
    ]
    pairs = verify_825.extract_tq_benke_pairs(rows)
    # full-width paren normalised to half-width, whitespace stripped
    assert pairs == [("P002", "软件工程(男)")]


# --- core contract: overlap report ----------------------------------------

def test_report_overlap_when_j3_subset_of_tq_returns_zero_unique():
    j3_pairs = [("B002", "数学"), ("B003", "英语")]
    tq_pairs = [("B002", "数学"), ("B003", "英语"), ("B004", "物理")]
    report = verify_825.report_overlap(j3_pairs, tq_pairs)
    assert report.reported_count == 0
    assert report.unique_keys == []


def test_report_overlap_records_unique_keys_in_input_order():
    j3_pairs = [("B002", "数学"), ("B999", "护理"), ("B003", "英语")]
    tq_pairs = [("B002", "数学"), ("B003", "英语")]
    report = verify_825.report_overlap(j3_pairs, tq_pairs)
    assert report.reported_count == 1
    assert report.unique_keys == [("B999", "护理")]


def test_report_overlap_dedupes_repeated_j3_keys():
    """A school may list the same major in multiple近三年 rows; the unique-key
    set must collapse duplicates so the reported count is the de-duplicated size."""
    j3_pairs = [
        ("B002", "数学"), ("B002", "数学"),  # duplicate
        ("B999", "护理"),
    ]
    tq_pairs = [("B002", "数学")]
    report = verify_825.report_overlap(j3_pairs, tq_pairs)
    assert report.reported_count == 1
    assert report.unique_keys == [("B999", "护理")]


# --- csv writer for human review ------------------------------------------

def test_write_unique_csv_reports_correct_row_count(tmp_path: Path):
    j3_pairs = [("B002", "数学"), ("B999", "护理")]
    tq_pairs = [("B002", "数学")]
    report = verify_825.report_overlap(j3_pairs, tq_pairs)
    out = tmp_path / "j3_early_only.csv"
    written = verify_825.write_unique_csv(report, out)
    # contract: returns the count it wrote, and that equals reported_count
    assert written == report.reported_count == 1
    with out.open(encoding="utf-8") as fh:
        records = list(csv.DictReader(fh))
    assert len(records) == 1
    assert records[0]["schoolcode"] == "B999"
    assert records[0]["majorname"] == "护理"


def test_write_unique_csv_with_zero_rows_still_writes_header(tmp_path: Path):
    report = verify_825.report_overlap([("B002", "数学")], [("B002", "数学")])
    out = tmp_path / "empty.csv"
    verify_825.write_unique_csv(report, out)
    with out.open(encoding="utf-8") as fh:
        records = list(csv.DictReader(fh))
    assert records == []


# --- real-workbook smoke (contract: not hard FAIL) ------------------------

class TestVerify825Smoke:
    """Smoke: real workbooks. Per Plan v2 this is a **契约测试 (不硬 FAIL)** —
    the contract is `reported_count == len(csv 独有行)`, surfaced for human
    review. The reported_count itself is reported, not asserted to be 0."""

    def test_smoke_real_overlap_report_and_csv(self, repo_root: Path, tmp_path: Path):
        import openpyxl

        wb_j3 = openpyxl.load_workbook(
            repo_root / "data" / "近三年学校批次专业线差统计.xlsx",
            read_only=True, data_only=True,
        )
        rows_j3 = list(wb_j3["统计结果"].iter_rows(values_only=True))
        wb_j3.close()
        wb_tq = openpyxl.load_workbook(
            repo_root / "data" / "山东省高考提前批录取数据.xlsx",
            read_only=True, data_only=True,
        )
        rows_tq = list(wb_tq[wb_tq.sheetnames[0]].iter_rows(values_only=True))
        wb_tq.close()

        j3_pairs = verify_825.extract_j3_early_pairs(rows_j3)
        tq_pairs = verify_825.extract_tq_benke_pairs(rows_tq)
        report = verify_825.report_overlap(j3_pairs, tq_pairs)

        # contract: reported_count reflects the actual unique-row count
        # written to the review CSV.
        out = tmp_path / "s2_j3_early_only.csv"
        written = verify_825.write_unique_csv(report, out)
        with out.open(encoding="utf-8") as fh:
            csv_rows = list(csv.DictReader(fh))

        assert written == report.reported_count
        assert written == len(csv_rows)
        # Every reported key is genuinely absent from the supplement本科 pool.
        tq_set = set(tq_pairs)
        for record in csv_rows:
            key = (record["schoolcode"], record["majorname"])
            assert key not in tq_set
