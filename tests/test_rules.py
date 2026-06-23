"""Tests for the three locked rules (spec §3, §5.2 element 6, §5.4).

Slice 3 Task 3.2 — three rule fixtures:
  1. 专科排除: 小标题含「专科」的 181 行全部排除; 常规批两侧 0 专科残留.
  2. 选科非差异化: 跨年政策漂移 (物理 vs 物理和化学) 不阻断匹配; 命中后当选科
     不一致时日志附加「选科政策漂移，已忽略」. 选科不进任何匹配键.
  3. 招生类别差异化: 普通 vs 中外合作 (同核心名) -> 不匹配 (不同轨道).

The 专科 and 招生类别 rules are already implemented in Slice 1/2 builders; this
slice adds tests that *lock* the contracts so future changes cannot regress
them. The 选科 drift log is implemented in stage1_5_coarse (Slice 3).
"""

from __future__ import annotations

from pathlib import Path


from scripts import stage0_merge, stage1_5_coarse, stage1_strict
from scripts.models import DaglubenRow, HistoryRow


def _hist(**kw) -> HistoryRow:
    base: HistoryRow = dict(  # type: ignore[assignment]
        school="", school_cat="", major="", stripped="", core="",
        subject="", J=None, T=None, source_table="常规批一段线",
    )
    base.update(kw)  # type: ignore[arg-type]
    return base


def _dl(**kw) -> DaglubenRow:
    base: DaglubenRow = dict(  # type: ignore[assignment]
        school="", school_cat="普通计划", major="", stripped="", core="",
        subject="", batch="4.常规批", src_row_idx=0,
    )
    base.update(kw)  # type: ignore[arg-type]
    return base


def _dl_row(row: list) -> list:
    """Pad a dagluben-shaped row to 12 cols."""
    width = 12
    return list(row) + [None] * (width - len(row))


# === Rule 1: 专科 排除 (RED — lock 181 rows, both pools clean) ===============

def test_build_dagluben_regular_excludes_zhuanke_subtitle_major_rows():
    """A 专科 subtitle must exclude the major row under it even if 代号/名称
    are both non-empty."""
    rows = [
        _dl_row(
            ["批次", "小标题", "学校代码", "学校名", "代号", "名称",
             "选科", "学制", "计划数", "备注", "年收费", "整行校准"]
        ),
        _dl_row(["4.常规批", "普通计划", "A001", "X大学", "01", "英语",
                 "不限", "4", "1"]),
        _dl_row(["4.常规批", "定向培养军士生(专科)", "C001", "D职业学院",
                 "03", "护理", "不限", "3", "50"]),
    ]
    out = stage0_merge.build_dagluben_regular(rows)
    assert len(out) == 1
    assert out[0]["school"] == "X大学"
    # No 专科残留 in the regular pool.
    assert all("专科" not in (r.get("school_cat", "") or "") for r in out)


def test_build_dagluben_early_excludes_zhuanke_subtitle_major_rows():
    """The same专科 exclusion applies to提前批 B类 (181 rows in real data)."""
    rows = [
        _dl_row(
            ["批次", "小标题", "学校代码", "学校名", "代号", "名称"]
        ),
        _dl_row(["2.提前批B类", "定向培养军士生(专科)", "C001", "D职业学院",
                 "03", "护理", "不限", "3", "50"]),
        _dl_row(["2.提前批B类", "公安政法类", "P010", "公安大学", "01", "治安学",
                 "历史", "4", "1"]),
    ]
    out = stage0_merge.build_dagluben_early(rows)
    assert len(out) == 1
    assert out[0]["school"] == "公安大学"


class TestZhuankeSmoke:
    """Smoke: real-data 专科 exclusion cardinalities (NOT RED but locking the
    181-row contract so a future regression cannot silently drop it)."""

    def test_smoke_zhuanke_excludes_181_major_rows(self, repo_root: Path):
        import openpyxl

        wb = openpyxl.load_workbook(
            repo_root / "data" / "山东省2026年大绿本招生计划.xlsx",
            read_only=True, data_only=True,
        )
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
        wb.close()

        # Count 专科-subtitle major rows directly from source.
        zhuanke_major = 0
        for i, row in enumerate(rows, start=1):
            if i == 1:
                continue
            subtitle = row[1] if len(row) > 1 else None
            code = row[4] if len(row) > 4 else None
            name = row[5] if len(row) > 5 else None
            is_major = code not in (None, "") and name not in (None, "")
            if subtitle and "专科" in str(subtitle) and is_major:
                zhuanke_major += 1
        assert zhuanke_major == 181

    def test_smoke_regular_and_early_pools_have_zero_zhuanke_residue(
        self, repo_root: Path
    ):
        import openpyxl

        wb = openpyxl.load_workbook(
            repo_root / "data" / "山东省2026年大绿本招生计划.xlsx",
            read_only=True, data_only=True,
        )
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
        wb.close()

        regular = stage0_merge.build_dagluben_regular(rows)
        early = stage0_merge.build_dagluben_early(rows)
        assert all("专科" not in (r.get("school_cat", "") or "") for r in regular)
        assert all("专科" not in (r.get("school_cat", "") or "") for r in early)

        # The 181 rows must be absent from the combined 本科 pool: 23887 + 1585
        # (regular + early本科) — none of them carry 专科.
        combined = len(regular) + len(early)
        assert combined == 23887 + 1585


# === Rule 2: 选科 non-differentiation (RED — drift does not block match) =====

def test_stage1_strict_ignores_subject_in_key():
    """Stage 1 key is (school, cat, stripped); 选科 is NOT in the key. Two
    rows with different 选科 but identical key still match."""
    history = [
        _hist(school="D大学", school_cat="", stripped="数学",
              subject="物理", J=60.0),
    ]
    dagluben = [
        _dl(school="D大学", school_cat="普通计划", stripped="数学",
            major="数学", core="数学", subject="物理和化学", src_row_idx=1),
    ]
    results = stage1_strict.match_strict(dagluben, history)
    assert results[0]["matched"] is True
    assert results[0]["J"] == 60.0


def test_stage1_5_logs_subject_drift_when_subjects_differ():
    """Spec §9 / §5.4: when a coarse match pairs rows whose 选科 differ, the
    log must append「选科政策漂移，已忽略」so a human reviewer can see it."""
    core_idx = stage1_5_coarse.build_core_idx(
        [_hist(school="D大学", school_cat="", core="数学", subject="物理", J=60.0)]
    )
    unmatched = [
        _dl(school="D大学", school_cat="普通计划",
            major="数学", core="数学", subject="物理和化学", src_row_idx=5),
    ]
    accepted, _ = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 1
    assert "选科政策漂移，已忽略" in accepted[0]["log"]


def test_stage1_5_multi_value_subject_drift_also_logged():
    """近三年 subject is multi-value ~37.5% (e.g.「物理 | 物理和化学」). Treat
    any inequality (including multi-value) as drift and log it."""
    core_idx = stage1_5_coarse.build_core_idx(
        [_hist(school="E大学", school_cat="", core="化学",
               subject="物理|物理和化学", J=55.0)]
    )
    unmatched = [
        _dl(school="E大学", school_cat="普通计划",
            major="化学", core="化学", subject="化学", src_row_idx=6),
    ]
    accepted, _ = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 1
    assert "选科政策漂移，已忽略" in accepted[0]["log"]


# === Rule 3: 招生类别 differentiation (RED — different tracks never match) ====

def test_stage1_strict_different_category_blocks_match():
    """普通 vs 中外合作 (same school, same stripped) -> different key -> miss."""
    history = [
        _hist(school="S大学", school_cat="中外合作办学",
              stripped="英语", J=40.0),
    ]
    dagluben = [
        _dl(school="S大学", school_cat="普通计划", stripped="英语",
            major="英语", core="英语", src_row_idx=1),
    ]
    results = stage1_strict.match_strict(dagluben, history)
    assert results[0]["matched"] is False


def test_stage1_5_different_category_same_core_never_matches():
    """Stage 1.5 buckets by (school, normalise_cat, core); 普通 vs 中外合作
    same core -> NOT in the same bucket -> no candidate -> still unmatched."""
    core_idx = stage1_5_coarse.build_core_idx(
        [_hist(school="S大学", school_cat="中外合作办学", core="英语", J=40.0)]
    )
    unmatched = [
        _dl(school="S大学", school_cat="普通计划",
            major="英语", core="英语", src_row_idx=2),
    ]
    accepted, still = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 0
    assert len(still) == 1


def test_stage1_5_same_non_default_category_can_match():
    """Both sides with the same non-default category (e.g. 中外合作) share a
    bucket -> coarse match proceeds normally."""
    core_idx = stage1_5_coarse.build_core_idx(
        [_hist(school="S大学", school_cat="中外合作办学",
               core="英语", subject="物理", J=40.0)]
    )
    unmatched = [
        _dl(school="S大学", school_cat="中外合作办学",
            major="英语(师范)", core="英语", subject="物理和化学", src_row_idx=3),
    ]
    accepted, _ = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 1
    assert accepted[0]["J"] == 40.0
