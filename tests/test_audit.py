"""Slice C — data-quality audit hard gate (spec V5-3, Plan v2 Slice C 修订).

iteration-3 (structured-columns, Plan v2 CRITICAL扩范围): the audit now keys
every log-dependent check off the structured「匹配方式」column **by name**
(not the legacy log string, not a column index). This makes the gate robust
to column re-ordering and removes the duplicated JUDGMENTAL_LOG_PREFIXES.

Check mapping (iteration-3):
  0  复核覆盖完备性 (judgmental coverage): every row whose 匹配方式 ∈
     {核心名匹配, agent 语义匹配} in the hierarchical output must appear in
     ``verify_*_result.jsonl`` with verdict=确定.
  1  每本科专业行「匹配方式」非空 (0 缺失).
  2  每张产出表 0 个全空数据行.
  3  字段映射回归: each produced table non-empty.
  4  随机 ≥30 匹配行 J/T 与近三年原值逐行比对 (精度区分): rows with 匹配方式 ∈
     {严格匹配, 核心名匹配, agent 语义匹配} compared to source history; 匹配方式 ==
     新增专业 compared to the estimate table.
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from scripts.audit_output import AuditReport, audit, main

# --- fixtures ---------------------------------------------------------------

# iteration-3: row-end = 7 cols (J/T + 5 structured). The audit reads
# 匹配方式 by name (not index), so we keep the column order aligned with
# write_outputs for readability.
HEADER = [
    "批次",
    "小标题",
    "学校代码",
    "学校名",
    "代号",
    "名称",
    "选考科目要求",
    "学制",
    "计划数",
    "学校备注",
    "年收费",
    "整行校准",
    "近三年统计线差",
    "近三年线差标准差",
    "匹配方式",
    "仅一年数据",
    "选科要求跨年变化",
    "二次复核",
    "原因说明",
]


def _write_hier(path: Path, rows: list[list]) -> None:
    """Write a hierarchical-style workbook (header + major rows)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(HEADER))
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    wb.close()


def _write_edge(path: Path, header: list[str], rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    wb.close()


def _good_output_dir(tmp_path: Path) -> Path:
    """A clean output dir that passes every check.

    Hierarchical / flat carry three 本科 major rows:
      - 严格匹配 (src_row_idx=2)
      - agent 语义匹配 (src_row_idx=3) — judgmental, confirmed in the verify jsonl
      - 新增专业 (src_row_idx=4) — estimate row; J/T from 今年新增往年没有的专业.xlsx
    Each row carries the 5 structured columns directly (no legacy log cell).
    """
    out = tmp_path / "output"
    out.mkdir()
    hier_rows = [
        # strict matched (src_row_idx=2)
        [
            "4.常规批",
            "普通计划",
            "A01",
            "示例大学",
            "01",
            "计算机",
            "物理",
            "4",
            "2",
            "",
            "",
            "01计算机",
            60.0,
            5.0,
            "严格匹配",
            "",
            "",
            "",
            "归一化专业名+招生类别一致",
        ],
        # agent 语义匹配 (src_row_idx=3) — judgmental, confirmed in verify jsonl
        [
            "4.常规批",
            "普通计划",
            "A01",
            "示例大学",
            "02",
            "数学",
            "物理",
            "4",
            "1",
            "",
            "",
            "02数学",
            70.0,
            None,
            "agent 语义匹配",
            "",
            "",
            "确定",
            "语义匹配：方向对齐",
        ],
        # new major estimate (src_row_idx=4) — J/T from estimate, rounded
        [
            "4.常规批",
            "普通计划",
            "A01",
            "示例大学",
            "03",
            "新专业",
            "物理",
            "4",
            "1",
            "",
            "",
            "03新专业",
            80.0,
            4.0,
            "新增专业",
            "",
            "",
            "",
            "估算=同校同选科(2)均值=80.0",
        ],
    ]
    _write_hier(out / "大绿本_完整版_含线差.xlsx", hier_rows)
    _write_hier(out / "大绿本_专业列表_含线差.xlsx", hier_rows)
    _write_edge(
        out / "今年新增往年没有的专业.xlsx",
        [
            "学校",
            "专业",
            "选科",
            "统计线差估算",
            "线差标准差估算",
            "退化级别",
            "样本量",
            "日志",
        ],
        [
            [
                "示例大学",
                "新专业",
                "物理",
                80.0,
                4.0,
                0,
                2,
                "新增专业：估算=同校同选科(2)均值=80.0",
            ]
        ],
    )
    _write_edge(
        out / "未能匹配的专业.xlsx",
        ["src_row_idx", "学校", "招生类别", "专业", "核心名", "选科", "批次", "日志"],
        [[5, "他大学", "", "其他", "其他", "物理", "4.常规批", "无法匹配：剩余未归类"]],
    )
    _write_edge(
        out / "往年有但今年停招的专业.xlsx",
        ["学校", "招生类别", "专业", "近三年统计线差", "近三年线差标准差", "日志"],
        [["他大学", "", "旧专业", 40.0, 3.0, "近三年有、2026 大绿本无"]],
    )
    _write_edge(
        out / "学校改名表.xlsx",
        ["2026新校名", "候选旧校名", "置信度", "2026本科专业数", "备注", "人工已核验"],
        [["新校", "旧校", 0.9, 5, "前身旧校", False]],
    )
    _write_edge(
        out / "今年新招生的学校.xlsx",
        ["2026新校名", "2026本科专业数", "日志"],
        [["全新校", 3, "2026 新增校，近三年无招生"]],
    )
    _write_edge(
        out / "往年有今年停招的学校.xlsx",
        ["历史旧校名", "日志"],
        [["消失校", "学校未在 2026 招生"]],
    )
    return out


def _good_intermediate_dir(tmp_path: Path) -> Path:
    inter = tmp_path / "intermediate"
    inter.mkdir()
    return inter


def _good_semantic_dir(tmp_path: Path, confirmed_idx: list[int]) -> Path:
    sem = tmp_path / "semantic-match"
    sem.mkdir()
    res = sem / "verify_batch_01_result.jsonl"
    lines = [
        json.dumps({"src_row_idx": i, "verdict": "确定", "reason": "核心名一致"})
        for i in confirmed_idx
    ]
    res.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return sem


def _good_data_dir(tmp_path: Path) -> Path:
    data = tmp_path / "data"
    data.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "统计结果"
    ws.append(
        [
            "批次",
            "校码",
            "校名",
            "专业名",
            "选科",
            "备注",
            "基础专业",
            "是否括号",
            "括号",
            "统计线差",
            "2023",
            "2024",
            "2025",
            "",
            "",
            "",
            "年数",
            "",
            "",
            "线差标准差",
        ]
    )
    ws.append(
        [
            "常规批一段线",
            "A01",
            "示例大学",
            "计算机",
            "物理",
            "",
            "计算机",
            "否",
            "",
            60.0,
            60.0,
            60.0,
            60.0,
            "",
            "",
            "",
            3,
            "",
            "",
            5.0,
        ]
    )
    ws.append(
        [
            "常规批一段线",
            "A01",
            "示例大学",
            "数学",
            "物理",
            "",
            "数学",
            "否",
            "",
            70.0,
            70.0,
            70.0,
            70.0,
            "",
            "",
            "",
            3,
            "",
            "",
            None,
        ]
    )
    wb.save(data / "近三年学校批次专业线差统计.xlsx")
    wb.close()
    return data


# --- check 0: judgmental coverage (keys off 匹配方式) ----------------------


def test_audit_passes_on_good_output(tmp_path: Path) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)

    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    assert isinstance(report, AuditReport)
    assert report.ok is True, report.checks


def test_check0_fails_when_judgmental_row_not_in_verify(tmp_path: Path) -> None:
    """A row with 匹配方式=核心名匹配 whose src_row_idx is absent from
    verify_*_result.jsonl → check 0 fails → ok=False."""
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[99])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)

    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    assert report.ok is False
    c0 = next(c for c in report.checks if c["name"].startswith("judgmental_coverage"))
    assert c0["passed"] is False
    assert "3" in c0["detail"]


def test_check0_fails_when_verify_jsonl_missing(tmp_path: Path) -> None:
    """No verify_*_result.jsonl present at all → fail with「复核未派发」."""
    out = _good_output_dir(tmp_path)
    sem = tmp_path / "semantic-match"
    sem.mkdir()
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)

    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    assert report.ok is False
    c0 = next(c for c in report.checks if c["name"].startswith("judgmental_coverage"))
    assert c0["passed"] is False
    assert "复核未派发" in c0["detail"]


def test_check0_does_not_treat_semantic_null_as_judgmental(tmp_path: Path) -> None:
    """A agent 语义匹配 row is judgmental; but a row with 匹配方式 NOT in
    {核心名匹配, agent 语义匹配} (e.g. 严格匹配 / 新增专业) is NOT judgmental — even
    if its note text happens to contain those keywords. Confirm the stage
    whitelist is the only discriminator."""
    out = _good_output_dir(tmp_path)
    # verify jsonl confirms only idx=3 (the 粗筛 row). The strict + 新增 rows
    # are NOT judgmental, so their absence from verify is fine → check0 passes.
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    c0 = next(c for c in report.checks if c["name"].startswith("judgmental_coverage"))
    assert c0["passed"] is True, c0


# --- check 1: every 本科 major row has non-empty 匹配方式 ----------------


def test_check1_fails_on_blank_stage(tmp_path: Path) -> None:
    """A flat major row with 匹配方式 blank → check 1 fails."""
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    # Corrupt the flat output: blank 匹配方式 on the first major row.
    wb = openpyxl.load_workbook(out / "大绿本_专业列表_含线差.xlsx")
    ws = wb.active
    # Find the 匹配方式 column by name (not index).
    header = [c.value for c in ws[1]]
    stage_col = header.index("匹配方式") + 1
    ws.cell(row=2, column=stage_col).value = None
    wb.save(out / "大绿本_专业列表_含线差.xlsx")
    wb.close()

    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    assert report.ok is False
    c1 = next(c for c in report.checks if c["name"].startswith("nonempty_log"))
    assert c1["passed"] is False


# --- check 2: no fully-empty data rows in any produced table ---------------


def test_check2_fails_on_empty_row(tmp_path: Path) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    wb = openpyxl.load_workbook(out / "未能匹配的专业.xlsx")
    ws = wb.active
    ws.append([None, None, None, None, None, None, None, None])
    wb.save(out / "未能匹配的专业.xlsx")
    wb.close()

    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    assert report.ok is False
    c2 = next(c for c in report.checks if c["name"].startswith("no_empty_rows"))
    assert c2["passed"] is False


# --- check 3: produced tables non-empty (field-mapping regression guard) ---


def test_check3_fails_when_table_has_no_data_rows(tmp_path: Path) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    _write_edge(
        out / "今年新增往年没有的专业.xlsx",
        [
            "学校",
            "专业",
            "选科",
            "统计线差估算",
            "线差标准差估算",
            "退化级别",
            "样本量",
            "日志",
        ],
        [],
    )

    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    assert report.ok is False
    c3 = next(c for c in report.checks if c["name"].startswith("tables_nonempty"))
    assert c3["passed"] is False


# --- check 4: J/T consistency (keys off 匹配方式 for matched/estimate) -----


def test_check4_fails_on_jt_mismatch(tmp_path: Path) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    history = [
        {"school": "示例大学", "major": "计算机", "J": 60.0, "T": 5.0},
        {"school": "示例大学", "major": "数学", "J": 70.0, "T": None},
    ]
    # Corrupt the strict row's J (匹配方式=严格匹配 → matched).
    wb = openpyxl.load_workbook(out / "大绿本_完整版_含线差.xlsx")
    ws = wb.active
    ws.cell(row=2, column=13).value = 999.0
    wb.save(out / "大绿本_完整版_含线差.xlsx")
    wb.close()

    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=history,
        semantic_dir=sem,
    )
    assert report.ok is False
    c4 = next(c for c in report.checks if c["name"].startswith("jt_consistency"))
    assert c4["passed"] is False


def test_check4_estimate_row_uses_round_tolerance(tmp_path: Path) -> None:
    """新增专业 rows (匹配方式=新增专业) compared against 今年新增往年没有的专业.xlsx."""
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    history = [
        {"school": "示例大学", "major": "计算机", "J": 60.0, "T": 5.0},
        {"school": "示例大学", "major": "数学", "J": 70.0, "T": None},
    ]
    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=history,
        semantic_dir=sem,
    )
    c4 = next(c for c in report.checks if c["name"].startswith("jt_consistency"))
    assert c4["passed"] is True, c4


# --- AuditReport structure + sample artefact -------------------------------


def test_audit_report_structure(tmp_path: Path) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    report = audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    assert hasattr(report, "ok")
    assert hasattr(report, "checks")
    assert isinstance(report.checks, list)
    assert len(report.checks) >= 5
    for c in report.checks:
        assert {"name", "passed", "detail"} <= set(c.keys())


def test_audit_writes_sample_xlsx(tmp_path: Path) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    audit(
        out,
        data_dir=data,
        intermediate_dir=inter,
        history=None,
        semantic_dir=sem,
    )
    sample = out / "audit_sample.xlsx"
    assert sample.exists()
    wb = openpyxl.load_workbook(sample, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    assert len(rows) >= 2


# --- main() exit code contract ---------------------------------------------


def test_main_exits_zero_on_pass(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "audit_output",
            "--output-dir",
            str(out),
            "--data-dir",
            str(data),
            "--intermediate-dir",
            str(inter),
            "--semantic-dir",
            str(sem),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 0


def test_main_exits_nonzero_on_fail(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    out = _good_output_dir(tmp_path)
    sem = tmp_path / "semantic-match"
    sem.mkdir()
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "audit_output",
            "--output-dir",
            str(out),
            "--data-dir",
            str(data),
            "--intermediate-dir",
            str(inter),
            "--semantic-dir",
            str(sem),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code != 0


def test_main_cli_subprocess_zero(tmp_path: Path) -> None:
    out = _good_output_dir(tmp_path)
    sem = _good_semantic_dir(tmp_path, confirmed_idx=[3])
    data = _good_data_dir(tmp_path)
    inter = _good_intermediate_dir(tmp_path)
    proc = subprocess.run(
        [
            sys.executable,
            "-m",
            "scripts.audit_output",
            "--output-dir",
            str(out),
            "--data-dir",
            str(data),
            "--intermediate-dir",
            str(inter),
            "--semantic-dir",
            str(sem),
        ],
        capture_output=True,
        text=True,
    )
    assert proc.returncode == 0, proc.stderr
