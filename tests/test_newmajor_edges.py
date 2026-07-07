"""Pure-function TDD for Slice 5 Task 5.2 — new-major edge-table writers.

Covers the functions in scripts/write_edge_tables.py:
    identify_new_majors(unmatched, history) -> list[DaglubenRow]
    write_new_major_table(new_majors_with_estimate, out_path) -> None

``mark_newmajor_in_main`` was removed in iteration-2 Slice D (issue #13) as
dead code — run_pipeline's ``_build_main_results`` reads the estimate dict
directly, the marker was never on the call path.

Small-sample RED cases only; the real-data smoke count is reported in
scripts/run_newmajor_smoke.py and does NOT participate in the RED contract.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from scripts.models import DaglubenRow, HistoryRow
from scripts.write_edge_tables import (
    identify_new_majors,
    write_new_major_table,
)


# ---------------------------------------------------------------------------
# identify_new_majors
# ---------------------------------------------------------------------------


def test_identify_new_majors_keeps_rows_with_no_core_match_in_school() -> None:
    # unmatched 大绿本专业；同校历史里无任何 core 名匹配 → 真·新增。
    unmatched = [
        DaglubenRow(
            school="示例大学", core="人工智能", major="人工智能", src_row_idx=5
        ),
        DaglubenRow(
            school="示例大学", core="量子信息", major="量子信息", src_row_idx=6
        ),
    ]
    history = [
        HistoryRow(school="示例大学", core="计算机", major="计算机"),
        HistoryRow(school="示例大学", core="英语", major="英语"),
    ]
    result = identify_new_majors(unmatched, history)
    assert [r["core"] for r in result] == ["人工智能", "量子信息"]


def test_identify_new_majors_excludes_rows_with_school_core_candidate() -> None:
    # 同校历史里存在同 core 名 → 不是真新增（归改名/无候选/特殊处理，不在本函数）。
    unmatched = [
        DaglubenRow(
            school="示例大学", core="人工智能", major="人工智能", src_row_idx=5
        ),
        DaglubenRow(
            school="示例大学", core="计算机", major="计算机(新方向)", src_row_idx=7
        ),
    ]
    history = [
        HistoryRow(school="示例大学", core="计算机", major="计算机"),
    ]
    result = identify_new_majors(unmatched, history)
    assert [r["core"] for r in result] == ["人工智能"]


def test_identify_new_majors_school_scoped_not_global() -> None:
    # core 名在 *别的学校* 出现不算同校候选 → 仍是新增。
    unmatched = [
        DaglubenRow(school="甲大学", core="人工智能", major="人工智能", src_row_idx=1),
    ]
    history = [
        HistoryRow(school="乙大学", core="人工智能", major="人工智能"),
    ]
    result = identify_new_majors(unmatched, history)
    assert len(result) == 1


def test_identify_new_majors_empty_unmatched_returns_empty() -> None:
    assert identify_new_majors([], [HistoryRow(school="x", core="y")]) == []


# ---------------------------------------------------------------------------
# write_new_major_table
# ---------------------------------------------------------------------------


def test_write_new_major_table_writes_estimate_and_level_and_log(
    tmp_path: Path,
) -> None:
    # V5-1: the table also carries the 线差标准差估算 column (T), sourced from
    # the record's "T" key (EstimateResult["T"]). It may be None (level 2 or
    # no compatible T) but the column and value must round-trip.
    rows = [
        {
            "school": "示例大学",
            "major": "人工智能",
            "subject": "物理和化学",
            "value": 80.0,
            "T": 12.5,
            "level": 0,
            "n": 2,
            "log": "新增专业：估算=同校同选科(2)均值=80.0",
        },
        {
            "school": "全新大学",
            "major": "量子信息",
            "subject": "物理和化学",
            "value": None,
            "T": None,
            "level": 2,
            "n": 0,
            "log": "新校/无历史，无法估算",
        },
    ]
    out_path = tmp_path / "今年新增往年没有的专业.xlsx"
    write_new_major_table(rows, out_path)

    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = all_rows[0]
    # 关键列必须存在 (V5-1: 线差标准差估算 新增).
    assert "统计线差估算" in header
    assert "线差标准差估算" in header
    assert "退化级别" in header
    assert "样本量" in header
    assert "日志" in header

    data = all_rows[1:]
    assert len(data) == 2
    assert data[0][header.index("统计线差估算")] == 80.0
    assert data[0][header.index("线差标准差估算")] == 12.5
    assert data[0][header.index("退化级别")] == 0
    assert data[1][header.index("统计线差估算")] is None
    assert data[1][header.index("线差标准差估算")] is None
    assert data[1][header.index("退化级别")] == 2


def test_write_new_major_table_empty_input_still_writes_header(
    tmp_path: Path,
) -> None:
    out_path = tmp_path / "今年新增往年没有的专业.xlsx"
    write_new_major_table([], out_path)
    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    assert len(rows) == 1  # header only
