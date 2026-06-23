"""Tests for scripts.stage1_5_coarse — Stage 1.5 core-name coarse match.

Per Plan v2 / spec §6 Stage 1.5 (prototype-validated): for Stage 1 misses,
key candidates by ``(基础校名, 招生类别, 核心名)``.
  - Unique candidate  -> auto-accept (log ``粗筛匹配：核心名唯一``).
  - Multi candidate   -> disambiguate by「近三年候选差异化括号 ⊂ 大绿本全名」
                         (性别/合作/其他 each must appear in dagluben全名);
                         exactly one compatible -> accept
                         (log ``粗筛匹配：括号子集消歧（<简述>）``); else unmatched.
  - No candidate      -> still unmatched.

Prototype实证: 大绿本 2026 名远比近三年详细, 核心名对齐即可; **禁用签名全等**.

Per Plan v2: small-sample 1-unique + 1-disambig + 1-no-cand + 1-still-ambiguous
is the RED判据; the ~74.4% cumulative rate is smoke (separate class, NOT RED).
"""

from __future__ import annotations

from pathlib import Path

import pytest

from scripts import stage0_merge, stage1_5_coarse, stage1_strict
from scripts.models import DaglubenRow, HistoryRow


def _hist(**kw) -> HistoryRow:
    base: HistoryRow = dict(  # type: ignore[assignment]
        school="", school_cat="", major="", stripped="", core="",
        subject="", J=None, T=None, source_table="常规批一段线",
    )
    base.update(kw)  # type: ignore[arg-type]
    return base


def _dl(**kw) -> DaglubenRow:
    base: DaglubenRow = dict(  # type: ignore[assignment]
        school="", school_cat="普通计划", major="", stripped="", core="",
        subject="", batch="4.常规批", src_row_idx=0,
    )
    base.update(kw)  # type: ignore[arg-type]
    return base


# --- build_core_idx (RED) ---------------------------------------------------

def test_build_core_idx_keys_by_school_cat_core():
    history = [
        _hist(school="中国人民大学", school_cat="", core="经济学类", J=70.0),
        _hist(school="中国人民大学", school_cat="", core="经济学类", J=60.0),
        _hist(school="北京大学", school_cat="", core="数学"),
    ]
    idx = stage1_5_coarse.build_core_idx(history)
    key = ("中国人民大学", "", "经济学类")
    assert key in idx
    assert len(idx[key]) == 2
    assert ("北京大学", "", "数学") in idx


def test_build_core_idx_folds_default_cat_like_stage1():
    """The普通计划 (dagluben subtitle) vs "" (history) default track must fold
    to the same key — Stage 1.5 reuses Stage 1's category normalisation so the
    coarse key is consistent with the strict key."""
    history = [_hist(school="Z大学", school_cat="", core="英语", J=50.0)]
    idx = stage1_5_coarse.build_core_idx(history)
    # dagluben 普通计划 -> normalise_cat -> "" must hit the same bucket.
    assert ("Z大学", "", "英语") in idx


# --- match_coarse: unique candidate -> auto-accept (RED, 实证样例) -----------

def test_match_coarse_unique_candidate_auto_accepts():
    """实证样例: 人大「经济学类(经济学、国民经济管理、…)」 core=经济学类,
    同校同类别仅 1 候选「经济学类」 -> 自动接受.

    日志 ``粗筛匹配：核心名唯一``.
    """
    core_idx = stage1_5_coarse.build_core_idx(
        [
            _hist(school="中国人民大学", school_cat="", major="经济学类",
                  core="经济学类", J=72.0, T=4.0),
        ]
    )
    unmatched = [
        _dl(school="中国人民大学", school_cat="普通计划",
            major="经济学类(经济学、国民经济管理、能源经济、国际经济与贸易)",
            core="经济学类", src_row_idx=10),
    ]
    accepted, still = stage1_5_coarse.match_coarse(unmatched, core_idx)

    assert len(accepted) == 1
    assert len(still) == 0
    r = accepted[0]
    assert r["matched"] is True
    assert r["src_row_idx"] == 10
    assert r["J"] == 72.0
    assert r["T"] == 4.0
    assert r["log"] == "粗筛匹配：核心名唯一"


def test_match_coarse_unique_candidate_does_not_require_full_name_equality():
    """Signature-equality is explicitly BANNED (prototype 0%). A far more
    detailed dagluben全名 matches a bare history core as long as the core
    name is the sole candidate for that (school, cat, core)."""
    core_idx = stage1_5_coarse.build_core_idx(
        [
            _hist(school="北航", school_cat="", major="数学与应用数学",
                  core="数学与应用数学", J=80.0, T=2.0),
        ]
    )
    unmatched = [
        _dl(school="北航", school_cat="普通计划",
            major="数学与应用数学(拔尖学生培养计划)(含双学位项目)",
            core="数学与应用数学", src_row_idx=7),
    ]
    accepted, _ = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 1
    assert accepted[0]["log"] == "粗筛匹配：核心名唯一"
    assert accepted[0]["J"] == 80.0


# --- match_coarse: multi-candidate bracket-subset disambiguation (RED) ------

def test_match_coarse_multi_candidate_disambiguates_when_brackets_subset():
    """实证样例: 北航「数学与应用数学(拔尖学生培养计划)(含…)」.
    Candidates: 「数学与应用数学(拔尖学生培养计划)」(only this one's bracket
    is a substring of the dagluben全名) and bare「数学与应用数学」.
    Bare candidate has no brackets -> trivially subset-compatible, so there are
    2 compatible candidates -> still ambiguous -> still unmatched.

    Constructed version below has exactly one compatible candidate so the
    bracket-subset path accepts it.
    """
    core_idx = stage1_5_coarse.build_core_idx(
        [
            _hist(school="北航", school_cat="",
                  major="数学与应用数学(拔尖学生培养计划)", core="数学与应用数学",
                  J=80.0, T=2.0),
            _hist(school="北航", school_cat="",
                  major="数学与应用数学(华罗庚班)", core="数学与应用数学",
                  J=90.0, T=1.0),
        ]
    )
    unmatched = [
        _dl(school="北航", school_cat="普通计划",
            major="数学与应用数学(拔尖学生培养计划)(含双学位项目)",
            core="数学与应用数学", src_row_idx=11),
    ]
    accepted, still = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 1
    assert len(still) == 0
    r = accepted[0]
    assert r["matched"] is True
    assert r["J"] == 80.0
    assert r["log"].startswith("粗筛匹配：括号子集消歧")


def test_match_coarse_multi_candidate_still_unmatched_when_ambiguous():
    """实证样例: 人大「计算机类(图灵…)」 多候选且括号非子集唯一 -> 仍 unmatched.

    Bare '计算机类' candidate (no brackets) is trivially subset-compatible;
    one of the named candidates also happens to match if its bracket is a
    substring. To cleanly demonstrate the「brackets非子集」path we construct
    two candidates whose brackets are NOT substrings of the dagluben全名 —
    then zero compatible -> still unmatched.
    """
    core_idx = stage1_5_coarse.build_core_idx(
        [
            _hist(school="中国人民大学", school_cat="",
                  major="计算机类(图灵实验班)", core="计算机类", J=70.0),
            _hist(school="中国人民大学", school_cat="",
                  major="计算机类(金融与大数据技术创新人才班)",
                  core="计算机类", J=60.0),
        ]
    )
    # dagluben全名 contains neither「图灵实验班」nor「金融与大数据技术创新人才班」.
    unmatched = [
        _dl(school="中国人民大学", school_cat="普通计划",
            major="计算机类(AI全栈工程师计划)",
            core="计算机类", src_row_idx=12),
    ]
    accepted, still = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 0
    assert len(still) == 1
    assert still[0]["matched"] is False


def test_match_coarse_multi_candidate_two_compatible_still_unmatched():
    """When two candidates are both subset-compatible the match is ambiguous
    and must fall through to Stage 2. Bare candidate (no brackets) + one
    bracketed candidate whose bracket IS a substring -> 2 compatible -> miss.
    """
    core_idx = stage1_5_coarse.build_core_idx(
        [
            _hist(school="X大学", school_cat="", major="英语", core="英语", J=50.0),
            _hist(school="X大学", school_cat="",
                  major="英语(师范)", core="英语", J=55.0),
        ]
    )
    unmatched = [
        _dl(school="X大学", school_cat="普通计划",
            major="英语(师范)", core="英语", src_row_idx=13),
    ]
    accepted, still = stage1_5_coarse.match_coarse(unmatched, core_idx)
    # bare「英语」(no brackets -> subset) + 「英语(师范)」 -> 2 compatible -> miss
    assert len(accepted) == 0
    assert len(still) == 1


# --- match_coarse: no candidate -> still unmatched (RED) --------------------

def test_match_coarse_no_candidate_leaves_unmatched():
    core_idx = stage1_5_coarse.build_core_idx(
        [_hist(school="北京大学", core="数学", J=50.0)]
    )
    unmatched = [
        _dl(school="北京大学", school_cat="普通计划",
            major="天体物理学", core="天体物理学", src_row_idx=9),
    ]
    accepted, still = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 0
    assert len(still) == 1
    assert still[0]["school"] == "北京大学"


# --- 选科 non-differentiation: subject never enters the coarse key (RED) ----

def test_match_coarse_subject_drift_does_not_block_match():
    """Spec §5.4: 选科 = 非差异化 (policy drift). Construct a dagluben row with
    选科「物理和化学」 whose unique candidate has 选科「物理」: still matches
    because 选科 is NOT part of the core key; log appends「选科政策漂移，已忽略」.
    """
    core_idx = stage1_5_coarse.build_core_idx(
        [
            _hist(school="D大学", school_cat="", core="数学",
                  subject="物理", J=60.0),
        ]
    )
    unmatched = [
        _dl(school="D大学", school_cat="普通计划",
            major="数学", core="数学", subject="物理和化学", src_row_idx=20),
    ]
    accepted, _ = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 1
    assert accepted[0]["matched"] is True
    assert "选科政策漂移，已忽略" in accepted[0]["log"]


def test_match_coarse_same_subject_no_drift_note():
    """When subjects agree, no drift note is appended."""
    core_idx = stage1_5_coarse.build_core_idx(
        [_hist(school="D大学", core="数学", subject="物理和化学", J=60.0)]
    )
    unmatched = [
        _dl(school="D大学", school_cat="普通计划",
            major="数学", core="数学", subject="物理和化学", src_row_idx=21),
    ]
    accepted, _ = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert "选科政策漂移" not in accepted[0]["log"]


# --- 招生类别 differentiation: different tracks never match via coarse (RED) -

def test_match_coarse_different_category_never_matches():
    """Spec §5.2 element 6: 招生类别 is a differentiator. 普通 vs 中外合作
    same core name -> NOT in the same core bucket -> no candidate -> miss.
    """
    core_idx = stage1_5_coarse.build_core_idx(
        [
            _hist(school="S大学", school_cat="中外合作办学",
                  major="英语", core="英语", J=40.0),
        ]
    )
    unmatched = [
        _dl(school="S大学", school_cat="普通计划",
            major="英语", core="英语", src_row_idx=22),
    ]
    accepted, still = stage1_5_coarse.match_coarse(unmatched, core_idx)
    assert len(accepted) == 0
    assert len(still) == 1


# --- combined Stage1 + Stage1.5 ordering (RED) ------------------------------

def test_stage1_then_stage1_5_pipeline_combines_results():
    """Stage1 hits + Stage1.5 auto-accepts must compose into a single ordered
    list with no overlap and no dropped rows."""
    history = [
        _hist(school="A大学", school_cat="", stripped="数学", core="数学", J=50.0),
        _hist(school="B大学", school_cat="", stripped="化学", core="化学", J=60.0),
        _hist(school="C大学", school_cat="", stripped="物理", core="物理", J=70.0),
    ]
    dagluben = [
        _dl(school="A大学", stripped="数学", major="数学", core="数学", src_row_idx=1),
        # Stage1 strict miss but Stage1.5 unique core -> auto-accept
        _dl(school="B大学", stripped="化学(含新能源)", major="化学(含新能源)",
            core="化学", src_row_idx=2),
        # No candidate at all
        _dl(school="Z大学", stripped="天文学", major="天文学", core="天文学",
            src_row_idx=3),
    ]
    strict_results = stage1_strict.match_strict(dagluben, history)

    # Run stage1.5 over the unmatched dagluben rows.
    unmatched_dl = [
        dagluben[i] for i, r in enumerate(strict_results) if not r["matched"]
    ]
    core_idx = stage1_5_coarse.build_core_idx(history)
    coarse_accepted, coarse_still = stage1_5_coarse.match_coarse(unmatched_dl, core_idx)

    assert len(coarse_accepted) == 1
    assert coarse_accepted[0]["src_row_idx"] == 2
    assert coarse_accepted[0]["J"] == 60.0
    assert len(coarse_still) == 1
    assert coarse_still[0]["src_row_idx"] == 3


# --- Real-workbook smoke: ~74.4% cumulative auto (NOT RED) ------------------

class TestStage1_5Smoke:
    """Smoke层: real regular-batch cumulative auto rate. Plan v2 binding:
    assert 72%-78% (prototype observed 74.4%). NOT part of RED."""

    def test_smoke_cumulative_auto_rate(self, repo_root: Path):
        import openpyxl

        wb_j3 = openpyxl.load_workbook(
            repo_root / "data" / "近三年学校批次专业线差统计.xlsx",
            read_only=True, data_only=True,
        )
        rows_j3 = list(wb_j3["统计结果"].iter_rows(values_only=True))
        wb_j3.close()
        wb_dl = openpyxl.load_workbook(
            repo_root / "data" / "山东省2026年大绿本招生计划.xlsx",
            read_only=True, data_only=True,
        )
        rows_dl = list(wb_dl[wb_dl.sheetnames[0]].iter_rows(values_only=True))
        wb_dl.close()

        hist = stage0_merge.build_history_regular(rows_j3)
        dl = stage0_merge.build_dagluben_regular(rows_dl)
        strict_results = stage1_strict.match_strict(dl, hist)

        unmatched_dl = [
            dl[i] for i, r in enumerate(strict_results) if not r["matched"]
        ]
        core_idx = stage1_5_coarse.build_core_idx(hist)
        accepted, _still = stage1_5_coarse.match_coarse(unmatched_dl, core_idx)

        strict_hits = len(dl) - len(unmatched_dl)
        total_auto = strict_hits + len(accepted)
        rate = total_auto / len(dl)
        # Plan: 74.4% observed; allow 72%-78% band.
        assert 0.72 <= rate <= 0.78, f"cumulative auto rate {rate:.4f} outside 72%-78%"
