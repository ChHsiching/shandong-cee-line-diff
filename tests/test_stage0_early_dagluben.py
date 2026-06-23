"""Tests for Stage 0 提前批 dagluben builder + Stage 1 提前批 coverage (Task 2.4).

Locks two contracts:
  1. ``build_dagluben_early`` merges大绿本 提前批 A类 + B类 本科专业 (excluding
     the 181 定向培养军士生(专科) rows in B类) into a single pool with a unified
     batch label, so Stage 1 can match it against the提前批 history pool.
  2. Stage 1 strict match, run over the full unified history + the merged
     提前批 dagluben pool, actually hits early-batch rows (not zero coverage).
"""

from __future__ import annotations

from pathlib import Path

import pytest

from scripts import stage0_merge, stage1_strict


# --- build_dagluben_early pure function (RED) -------------------------------

def _dl_row(row: list) -> list:
    """Pad to 12 cols (dagluben width)."""
    width = 12
    return list(row) + [None] * (width - len(row))


def test_build_dagluben_early_keeps_a_and_b_drops_zhuanke_subtitle():
    rows = [
        _dl_row(
            ["批次", "小标题", "学校代码", "学校名", "代号", "名称",
             "选科", "学制", "计划数", "备注", "年收费", "整行校准"]
        ),
        # A类 major row — kept
        _dl_row(["1.提前批A类", "军事类", "P002", "国防科技大学", "10", "数学",
                 "物理和化学", "4", "1"]),
        # B类 major row — kept (merged)
        _dl_row(["2.提前批B类", "公安政法类", "P010", "公安大学", "01", "治安学",
                 "历史", "4", "1"]),
        # B类 专科 subtitle major row — dropped
        _dl_row(["2.提前批B类", "定向培养军士生(专科)", "C001", "D职业学院",
                 "03", "护理", "不限", "3", "50"]),
        # 4.常规批 row — dropped (out of scope)
        _dl_row(["4.常规批", "普通计划", "A001", "X大学", "01", "英语",
                 "不限", "4", "1"]),
    ]
    out = stage0_merge.build_dagluben_early(rows)
    assert len(out) == 2
    assert {r["school"] for r in out} == {"国防科技大学", "公安大学"}
    # Both A and B fold into the unified batch label for the matching pool.
    assert {r["batch"] for r in out} == {"提前批"}


def test_build_dagluben_early_normalises_fields_like_regular():
    rows = [
        _dl_row(
            ["批次", "小标题", "学校代码", "学校名", "代号", "名称"]
        ),
        _dl_row(
            ["1.提前批A类", "军事类", "P002", "国防科技大学", "01",
             "软件工程（男）", "物理和化学", "4", "1"]
        ),
    ]
    out = stage0_merge.build_dagluben_early(rows)
    row = out[0]
    assert row["school"] == "国防科技大学"
    assert row["school_cat"] == "军事类"
    assert row["major"] == "软件工程(男)"          # nfk applied
    assert row["stripped"] == "软件工程(男)"       # gender preserved
    assert row["core"] == "软件工程"
    assert "src_row_idx" in row


# --- Real-workbook smoke (NOT RED) -----------------------------------------

class TestStage0EarlySmoke:
    def test_smoke_dagluben_early_row_count(self, repo_root: Path):
        """大绿本 提前批 本科专业 = A类 1139 + B类本科 446 = 1585
        (B类 181 定向培养军士生(专科) rows excluded per spec §3)."""
        import openpyxl

        wb = openpyxl.load_workbook(
            repo_root / "data" / "山东省2026年大绿本招生计划.xlsx",
            read_only=True, data_only=True,
        )
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
        wb.close()
        built = stage0_merge.build_dagluben_early(rows)
        assert len(built) == 1585


# --- Stage 1 covers early batch (RED contract) -----------------------------

def test_stage1_strict_hits_early_batch_rows():
    """Stage 1 over the unified history + merged early dagluben pool must
    produce non-zero early-batch matches. Synthetic mini-case."""
    history = [
        # A提前批 history row whose strict key matches the dagluben row below.
        {"school": "国防科技大学", "school_cat": "军事类",
         "stripped": "数学", "core": "数学", "major": "数学",
         "J": 100.0, "T": 5.0, "source_table": "提前批"},
    ]
    dagluben = [
        {"school": "国防科技大学", "school_cat": "军事类",
         "stripped": "数学", "core": "数学", "major": "数学",
         "subject": "物理和化学", "batch": "提前批", "src_row_idx": 2},
    ]
    results = stage1_strict.match_strict(dagluben, history)
    assert len(results) == 1
    assert results[0]["matched"] is True
    assert results[0]["J"] == 100.0
    assert results[0]["T"] == 5.0


class TestStage1EarlySmoke:
    """Smoke: real-data Stage1 coverage on提前批. Asserts a non-zero hit rate
    so the Slice 2 contract (Stage1 covers提前批) is locked without pinning
    a brittle percentage (Plan v2: rate is smoke, not RED)."""

    def test_smoke_stage1_early_batch_hit_count_nonzero(self, repo_root: Path):
        import openpyxl

        wb_j3 = openpyxl.load_workbook(
            repo_root / "data" / "近三年学校批次专业线差统计.xlsx",
            read_only=True, data_only=True,
        )
        rows_j3 = list(wb_j3["统计结果"].iter_rows(values_only=True))
        wb_j3.close()
        wb_tq = openpyxl.load_workbook(
            repo_root / "data" / "山东省高考提前批录取数据.xlsx",
            read_only=True, data_only=True,
        )
        rows_tq = list(wb_tq[wb_tq.sheetnames[0]].iter_rows(values_only=True))
        wb_tq.close()
        wb_dl = openpyxl.load_workbook(
            repo_root / "data" / "山东省2026年大绿本招生计划.xlsx",
            read_only=True, data_only=True,
        )
        rows_dl = list(wb_dl[wb_dl.sheetnames[0]].iter_rows(values_only=True))
        wb_dl.close()

        unified = stage0_merge.build_unified_history(rows_j3, rows_tq)
        early_pool = stage0_merge.build_dagluben_early(rows_dl)
        results = stage1_strict.match_strict(early_pool, unified)
        hits = sum(1 for r in results if r["matched"])
        assert hits > 0, "Stage1 must cover at least some early-batch rows"
        # Sanity: do not exceed the pool size.
        assert hits <= len(early_pool)
