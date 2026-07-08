"""Regression: output data-quality gaps (Phase 2 acceptance feedback).

Two bugs surfaced when the user reviewed the actual output:
- 本科 majors matched by none of strict/coarse/semantic/new/rename (特殊) got
  NO MatchResult → blank 匹配日志 in the output (spec requires a log on every row).
- 扁平版 included 专科 专业行 (scope is 仅本科)."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from scripts.constants import LOG_SPECIAL_UNMATCHED, LOG_ZHUANKE_OUT_OF_SCOPE
from scripts.models import DaglubenRow
from scripts.run_pipeline import _build_main_results
from scripts.write_outputs import write_flat


def test_build_main_results_emits_log_for_unmatched_special() -> None:
    """A 本科 row matched by none of strict/coarse/semantic/new/rename must still
    receive a MatchResult with a special-case log (was: dropped → blank 日志)."""
    dagluben = [
        DaglubenRow(
            src_row_idx=5,
            school="X大学",
            school_cat="普通计划",
            major="未知专业",
            batch="4.常规批",
        ),
    ]
    out = _build_main_results(dagluben, [], [], [], {})
    assert len(out) == 1
    assert out[0]["src_row_idx"] == 5
    assert out[0]["matched"] is False
    assert out[0]["J"] is None and out[0]["T"] is None
    assert out[0]["log"]
    assert LOG_SPECIAL_UNMATCHED in out[0]["log"]


def test_build_main_results_new_major_carries_estimate_T() -> None:
    """V5-1 / Plan v2 阻断1: a 新增专业 row's T must come from the estimate,
    not a hardcoded None. The estimate's value (J) and T are both surfaced."""
    from scripts.models import EstimateResult

    dagluben = [
        DaglubenRow(
            src_row_idx=7,
            school="Y大学",
            school_cat="普通计划",
            major="新专业Z",
            batch="4.常规批",
        ),
    ]
    estimates = {
        7: EstimateResult(value=88.0, T=13.5, level=0, n=2, log="估算log"),
    }
    out = _build_main_results(dagluben, [], [], [], estimates)
    assert len(out) == 1
    assert out[0]["src_row_idx"] == 7
    assert out[0]["J"] == 88.0
    assert out[0]["T"] == 13.5  # from estimate, NOT hardcoded None
    assert out[0]["log"] == "估算log"


def test_write_flat_excludes_zhuanke_major_rows(tmp_path: Path) -> None:
    """扁平版 must omit 专科 专业行 (小标题含「专科」); scope is 仅本科."""
    src = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "批次",
            "小标题",
            "学校代码",
            "学校名",
            "代号",
            "名称",
            "选科",
            "学制",
            "计划",
            "备注",
            "收费",
            "校准",
        ]
    )
    ws.append(
        [
            "4.常规批",
            "普通计划",
            "D001",
            "本大学",
            "01",
            "计算机",
            "物理和化学",
            "4",
            "10",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "2.提前批B类",
            "定向培养军士生(专科)",
            "D002",
            "专学院",
            "02",
            "电气技术",
            "物理",
            "3",
            "5",
            "",
            "",
            "",
        ]
    )
    wb.save(src)
    wb.close()

    out = tmp_path / "flat.xlsx"
    write_flat(src, [], out)

    ws2 = openpyxl.load_workbook(out, read_only=True).active
    rows = list(ws2.iter_rows(values_only=True))
    majors = [r for r in rows[1:] if r[4] not in (None, "")]
    assert len(majors) == 1, f"expected only the 本科 row, got {len(majors)}"
    assert majors[0][5] == "计算机"
    # 专科 row must not appear anywhere.
    assert all("专科" not in (r[1] or "") for r in rows)


def test_zhuanke_log_constants_exist() -> None:
    """Guard: the two new log constants are defined and non-empty."""
    assert LOG_SPECIAL_UNMATCHED
    assert LOG_ZHUANKE_OUT_OF_SCOPE


def test_build_dagluben_early_includes_flight_rows() -> None:
    """飞行技术(军队) 2 行 (batch 3) must enter the 提前批 pool (spec §6:
    归入提前批池匹配), not be excluded — otherwise they get no 匹配日志."""
    from scripts.constants import BATCH_EARLY_A, FLIGHT_BATCH
    from scripts.stage0_merge import build_dagluben_early

    rows = [
        [
            "批次",
            "小标题",
            "学校代码",
            "学校名",
            "代号",
            "名称",
            "选科",
            "学制",
            "计划",
            "备注",
            "收费",
            "校准",
        ],
        [
            BATCH_EARLY_A,
            "军事类",
            "P001",
            "军大",
            "01",
            "指挥",
            "物理和化学",
            "4",
            "5",
            "",
            "",
            "",
        ],
        [
            FLIGHT_BATCH,
            None,
            "P002",
            "飞大",
            "02",
            "飞行技术",
            "物理和化学",
            "4",
            "3",
            "",
            "",
            "",
        ],
    ]
    out = build_dagluben_early(rows)
    majors = [r["major"] for r in out]
    assert "飞行技术" in majors, f"flight row must be in pool, got {majors}"


def _load_xlsx(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    return rows


def test_write_deleted_major_table_maps_fields_to_columns(tmp_path: Path) -> None:
    """Regression (same class as 改名表): DeletedMajor field names must map to
    被删旧专业 header columns — passing field-named rows directly left cells empty."""
    from scripts.write_edge_tables import write_deleted_major_table

    rows = [
        {
            "school": "X大学",
            "school_cat": "普通计划",
            "major": "旧专业",
            "J": 50.0,
            "T": 5.0,
            "log": "近三年有、2026 大绿本无",
        }
    ]
    out = tmp_path / "被删.xlsx"
    write_deleted_major_table(rows, out)
    data = _load_xlsx(out)
    assert data[1][0] == "X大学"
    assert data[1][2] == "旧专业"
    assert data[1][3] == 50.0
    assert data[1][5] == "近三年有、2026 大绿本无"


def test_write_special_table_maps_fields_to_columns(tmp_path: Path) -> None:
    """Regression (same class): EdgeRow field names must map to 特殊情况 header.

    Task 3: 「对不上」行带 4 列估算（统计线差估算/线差标准差估算/估算方式/
    用了几条往年数据），插在 原因说明 前——飞行行这几列留空。"""
    from scripts.write_edge_tables import write_special_table

    rows = [
        {
            "src_row_idx": 7,
            "school": "飞大",
            "school_cat": "",
            "major": "飞行技术",
            "core": "飞行技术",
            "subject": "物理和化学",
            "batch": "提前批",
            "log": "飞行技术(军队)，提前批池匹配不成",
        },
        {
            "src_row_idx": 8,
            "school": "甲大学",
            "school_cat": "",
            "major": "工科试验班类",
            "core": "工科试验班类",
            "subject": "物理和化学",
            "batch": "4.常规批",
            "log": "没找到能匹配的往年专业：工科试验班类",
            "est_value": 82.5,
            "est_t": 3.2,
            "est_level": 0,
            "est_n": 12,
        },
    ]
    out = tmp_path / "特.xlsx"
    write_special_table(rows, out)
    data = _load_xlsx(out)
    header = data[0]
    assert "统计线差估算" in header
    assert "估算方式" in header
    assert header[-1] == "原因说明"
    j_idx = header.index("统计线差估算")
    log_idx = header.index("原因说明")
    way_idx = header.index("估算方式")

    flight = next(r for r in data[1:] if r[0] == 7)
    assert flight[log_idx] == "飞行技术(军队)，提前批池匹配不成"
    assert flight[j_idx] in (None, "")  # 飞行行不估

    other = next(r for r in data[1:] if r[0] == 8)
    assert other[j_idx] == 82.5
    assert other[way_idx] == "同校同选科均值"


# ---------------------------------------------------------------------------
# Slice D (issue #13) — field-mapping regression lock for ALL active writers.
# Each writer is asserted with **real field names** (the names the producing
# code actually passes in) so a field→header remap breakage fails loudly.
# Mirrors the陷阱 A that bit write_rename_table / write_special_table /
# write_deleted_major_table earlier.
# ---------------------------------------------------------------------------


def test_write_rename_table_maps_fields_to_columns(tmp_path: Path) -> None:
    """RenameRow fields (new_school/old_school/confidence/major_count_2026/
    remark/manual_reviewed) must reach the header columns, not vanish."""
    from scripts.write_edge_tables import write_rename_table

    rows = [
        {
            "new_school": "新大学",
            "old_school": "旧大学",
            "confidence": 0.88,
            "major_count_2026": 12,
            "remark": "网查：2026 由旧大学更名",
            "manual_reviewed": True,
        }
    ]
    out = tmp_path / "改名.xlsx"
    write_rename_table(rows, out)
    data = _load_xlsx(out)
    # header: 新校名 / 原校名 / 今年本科专业数 / 说明（含官方来源）
    assert data[1][0] == "新大学"
    assert data[1][1] == "旧大学"
    assert data[1][2] == 12
    assert data[1][3] == "网查：2026 由旧大学更名"


def test_write_new_school_table_maps_fields_to_columns(tmp_path: Path) -> None:
    """新增校表 records (new_school/major_count_2026) must populate cells."""
    from scripts.write_edge_tables import write_new_school_table

    rows = [{"new_school": "全新大学", "major_count_2026": 7}]
    out = tmp_path / "新增校.xlsx"
    write_new_school_table(rows, out)
    data = _load_xlsx(out)
    assert data[1][0] == "全新大学"
    assert data[1][1] == 7
    assert "新增校" in data[1][2]


def test_write_gone_school_table_maps_fields_to_columns(tmp_path: Path) -> None:
    """停招消失校表 records (old_school) must populate cells (not blank)."""
    from scripts.write_edge_tables import write_gone_school_table

    rows = [{"old_school": "消失大学"}]
    out = tmp_path / "停招.xlsx"
    write_gone_school_table(rows, out)
    data = _load_xlsx(out)
    assert data[1][0] == "消失大学"
    assert "未在 2026 招生" in data[1][1]


def test_write_new_major_table_maps_fields_to_columns(tmp_path: Path) -> None:
    """新增专业 table: record keys (school/major/subject/value/T/level/n/log)
    must reach the header columns — both J and T columns round-trip (V5-1)."""
    from scripts.write_edge_tables import write_new_major_table

    rows = [
        {
            "school": "新大学",
            "major": "人工智能",
            "subject": "物理和化学",
            "value": 88.5,
            "T": 6.25,
            "level": 0,
            "n": 3,
            "log": "新增专业：估算=同校同选科(3)均值=88.5",
        }
    ]
    out = tmp_path / "今年新增往年没有的专业.xlsx"
    write_new_major_table(rows, out)
    data = _load_xlsx(out)
    # header: 学校 / 专业 / 选科 / 统计线差估算 / 线差标准差估算 / 估算方式 / 用了几条往年数据 / 说明
    assert data[1][0] == "新大学"
    assert data[1][1] == "人工智能"
    assert data[1][2] == "物理和化学"
    assert data[1][3] == 88.5
    assert data[1][4] == 6.25
    assert data[1][5] == "同校同选科均值"  # level 0 → 大白话（#13）
    assert data[1][6] == 3
    assert "新增专业" in data[1][7]
